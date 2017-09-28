# -*- coding: utf-8 -*-
import os, django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ControlObras.settings")
django.setup()

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl import load_workbook
from datetime import datetime

from vista.models import ParqueSolar, Aerogenerador
from fu.models import ComponentesParque, Componente, EstadoFU, Registros

def getLastColumn(ws):
    fila = 3
    columna = 4
    valor = ws.cell(row=fila,column=columna).value
    while valor is not None:
        # Verifica que los nombres sean componentes válidos
        columna = columna + 1
        valor = ws.cell(row=fila, column=columna).value
    return columna

def checkValidFile(ws, componentes_parque):
    fila = 4
    columna = 2
    valor = ws.cell(row=fila,column=columna).value
    while valor is not None:
        elementos = componentes_parque.componentes.filter(nombre=valor)
        if elementos.count() <= 0:
            print valor
            return -1
        fila = fila +2
        valor = ws.cell(row=fila, column=columna).value

    return fila

if __name__ == '__main__':
    try:
        parque = ParqueSolar.objects.get(slug='lap-003')
    except ParqueSolar.DoesNotExist:
        parque = ParqueSolar.objects.get(slug='cli-001')

    print parque.nombre
    componentes_parque = ComponentesParque.objects.get(parque=parque)
    nombre_archivo = './registros.xlsx'
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    last_row = checkValidFile(ws, componentes_parque)
    if last_row == -1:
        # Retorna falso, no se pudo leer este archivo correctamente.
        print 'mal mal'
    else:
        print 'archivo válido'
    print last_row
    last_column = getLastColumn(ws)
    print last_column
    fila = 4
    ultimo_estado = None
    ultimo_anho = None
    estado = None
    while fila < last_row:
        columna = 4
        nombre_componente = ws.cell(row=fila, column=2).value
        print nombre_componente
        # lo puedo hacer de forma segura, porque se supone que los validé en checkValidFile
        componente = Componente.objects.get(nombre=nombre_componente)
        nombre_estado = ws.cell(row=fila, column=1).value

        if nombre_estado is not None:
            if nombre_estado != ultimo_estado:
                ultimo_estado = nombre_estado
                estado = None
                if ultimo_estado == 'Descarga en Parque':
                    estado = EstadoFU.objects.get(idx=1)
                elif ultimo_estado == 'Pre-montaje':
                    estado = EstadoFU.objects.get(idx=2)
                elif ultimo_estado == 'Montaje':
                    estado = EstadoFU.objects.get(idx=3)
                elif ultimo_estado == 'Puesta en marcha':
                    estado = EstadoFU.objects.get(idx=4)
                else:
                    estado = None
                    print('No existe el estado')
                    print nombre_estado
                print ultimo_estado

        if ultimo_estado is not None and estado is not None:
            while columna < last_column:
                aerogenerador_string = ws.cell(row=3, column=columna).value
                ag_num = aerogenerador_string.split(" ")[1]
                try:
                    aerogenerador = Aerogenerador.objects.get(parque=parque,nombre='WTG'+ag_num)
                except Aerogenerador.DoesNotExist:
                    print 'No existe...'
                fecha = ws.cell(row=fila, column=columna).value
                serie = ws.cell(row=fila + 1, column=columna).value

                if fecha is not None:
                    if serie is not None:
                        r = Registros(parque=parque,
                                      aerogenerador=aerogenerador,
                                      componente=componente,
                                      estado=estado,
                                      fecha=fecha.date(),
                                      no_serie=serie)
                        r.save()
                    else:
                        r = Registros(parque=parque,
                                      aerogenerador=aerogenerador,
                                      componente=componente,
                                      estado=estado,
                                      fecha=fecha.date())
                        r.save()

                columna += 1
        fila += 2
