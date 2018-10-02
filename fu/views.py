# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.conf import settings
import os
import StringIO
from collections import OrderedDict, defaultdict
from dateutil import relativedelta
import numpy as np
from anytree import Node, Resolver

from django.db.models import Sum
from django.core.exceptions import PermissionDenied
from django.core.urlresolvers import reverse
from django.template.response import TemplateResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.utils.timezone import localtime
from vista.models import ParqueSolar, Aerogenerador
from vista.functions import *
from fu.forms import ComponenteForm, AddComponentesForm, ConfiguracionFUForm,PlanificacionForm,DeleteComponentesForm
from fu.forms import RegistroDescargaForm, RegistroForm, ParadasForm,ReporteForm
from fu.models import Componente, ConfiguracionFU, Contractual, Plan, EstadoFU
from fu.models import ParqueEolico, Membership
from fu.models import Registros, Paradas
from ncr.views import serialize_grafico
from usuarios.models import Log
from querystring_parser import parser

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import SeriesLabel, StrRef
from openpyxl.drawing.fill import ColorChoice
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.text import Font as TextFont
from openpyxl.chart.text import RichText
from datetime import datetime
import base64
from easy_pdf.rendering import render_to_pdf
from django.core.exceptions import MultipleObjectsReturned
import json

meses_espanol = {"1": "Enero",
                 "2": "Febrero",
                 "3": "Marzo",
                 "4": "Abril",
                 "5": "Mayo",
                 "6": "Junio",
                 "7": "Julio",
                 "8": "Agosto",
                 "9": "Septiembre",
                 "10": "Octubre",
                 "11": "Noviembre",
                 "12": "Diciembre"
                 }


logger = logging.getLogger('oritec')


def getContractual(parque, componente, estado, fecha):
    c = Contractual.objects.filter(parque=parque,
                                   componente=componente,
                                   estado=estado,
                                   fecha__lte=fecha).aggregate(Sum('no_aerogeneradores'))
    if c['no_aerogeneradores__sum'] is None:
        return 0
    else:
        return c['no_aerogeneradores__sum']


def getPlan(parque, componente, estado, fecha):
    c = Plan.objects.filter(parque=parque,
                            componente=componente,
                            estado=estado,
                            fecha__lte=fecha).aggregate(Sum('no_aerogeneradores'))

    if c['no_aerogeneradores__sum'] is None:
        return 0
    else:
        return c['no_aerogeneradores__sum']


def getReal(parque, componente, estado, fecha):
    c = Registros.objects.filter(parque=parque,
                                 componente=componente,
                                 estado=estado,
                                 fecha__lte=fecha)
    return c.count()


def graficoComponentes(parque_eolico, estado, fecha_calculo):
    data_full = []
    # Me entrega el domingo final de esa semana.

    r = fecha_calculo
    max_aerogeneradores = parque_eolico.parque.no_aerogeneradores

    # Los totales
    data_graficos = []
    miembros = Membership.objects.filter(parque_eolico=parque_eolico, estado= estado).order_by('orden')
    for m in miembros:
        data_graficos.append({"name": m.componente.nombre, "y": max_aerogeneradores})
    data_full.append({"name": "Total", "data": data_graficos})

    # Contractual
    data_graficos = []

    for m in miembros:
        value = getContractual(parque_eolico.parque, m.componente, estado, r)
        data_graficos.append({"name": m.componente.nombre, "y": value})

    data_full.append({"name": "Contractual", "data": data_graficos})

    # Plan
    data_graficos = []
    for m in miembros:
        value = getPlan(parque_eolico.parque, m.componente, estado, r)
        data_graficos.append({"name": m.componente.nombre, "y": value})

    data_full.append({"name": "Plan", "data": data_graficos})

    # Real
    data_graficos = []
    for m in miembros:
        value = getReal(parque_eolico.parque, m.componente, estado, r)
        data_graficos.append({"name": m.componente.nombre, "y": value})
    data_full.append({"name": "Real", "data": data_graficos})

    datos = serialize_grafico(data_full)
    return datos


def calcularProyeccion(parque_eolico, anho, semana):
    d = str(anho) + '-W' + str(semana)
    # Se calcula solo con los datos de hasta la semana pasada
    r = datetime.strptime(d + '-0', "%Y-W%W-%w") - relativedelta.relativedelta(weeks=1)
    parque = parque_eolico.parque
    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return [[], None]
    max_aerogeneradores = parque.no_aerogeneradores
    estado_montaje = EstadoFU.objects.get(nombre='Montaje')
    ag_montar = aerogeneradoresAMontar(parque_eolico)
    # aux = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
    #                                   orden_montaje__range=(1, 8))
    aux = Membership.objects.filter(parque_eolico= parque_eolico,
                                    estado=estado_montaje,
                                    orden__range=(1, ag_montar))
    componentes_montaje = []
    for c in aux:
        componentes_montaje.append(c.componente.id)
    # Calculo de avance
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    # Valores a numpy
    x_values = []
    y_values = []
    first = False
    first_date = None
    count = 0

    while fecha_calculo <= r:
        c = Registros.objects.filter(parque=parque,
                                     componente__in=componentes_montaje,
                                     estado=estado_montaje,
                                     fecha__lte=fecha_calculo)
        valor = float(c.count()) / ag_montar
        if not first:
            if valor != 0:
                first = True
                first_date = fecha_calculo
        if first:
            if valor < max_aerogeneradores:
                x_values.append(count)
                y_values.append(valor)
                count += 1

        fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
        semana_calculo = fecha.isocalendar()[1]
        d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")

    if len(x_values) < 2:
        return [[], None]
    # Proyeccion
    x = np.array(x_values)
    y = np.array(y_values)
    z = np.polyfit(x, y, 1)
    p = np.poly1d(z)
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_graficos = []
    count = 0
    valor = 0
    while valor < max_aerogeneradores:
        if fecha_calculo < first_date:
            valor = 0
        else:
            valor = p(count)
            if valor < 0:
                valor = 0
            elif valor > max_aerogeneradores:
                valor = max_aerogeneradores
            count += 1
        fecha_grafico = str(fecha_calculo.year) + '-' + str(semana_calculo)
        valor_porcentaje = valor / max_aerogeneradores * 100
        data_graficos.append(valor_porcentaje)
        if valor < max_aerogeneradores:
            fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
            semana_calculo = fecha.isocalendar()[1]
            d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
            fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    return [data_graficos,fecha_calculo]


def calcularProyeccionGrafico(parque_eolico, anho, semana):
    d = str(anho) + '-W' + str(semana)
    # Se calcula solo con los datos de hasta la semana pasada
    r = datetime.strptime(d + '-0', "%Y-W%W-%w") - relativedelta.relativedelta(weeks=1)
    parque = parque_eolico.parque
    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return [[], None]
    max_aerogeneradores = parque.no_aerogeneradores
    # aux = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
    #                                   orden_montaje__range=(1, 8))
    estado_montaje = EstadoFU.objects.get(nombre='Montaje')

    ag_montar = aerogeneradoresAMontar(parque_eolico)

    aux = Membership.objects.filter(parque_eolico=parque_eolico,
                                    estado=estado_montaje,
                                    orden__range=(1,ag_montar))
    componentes_montaje = []
    for c in aux:
        componentes_montaje.append(c.componente.id)

    # Calculo de avance
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    # Valores a numpy
    x_values = []
    y_values = []
    first = False
    first_date = None
    count = 0

    while fecha_calculo <= r:
        c = Registros.objects.filter(parque=parque,
                                     componente__in=componentes_montaje,
                                     estado=estado_montaje,
                                     fecha__lte=fecha_calculo)
        valor = float(c.count()) / ag_montar
        if not first:
            if valor != 0:
                first = True
                first_date = fecha_calculo
        if first:
            if valor < max_aerogeneradores:
                x_values.append(count)
                y_values.append(valor)
                count += 1

        fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
        semana_calculo = fecha.isocalendar()[1]
        d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")

    if len(x_values) < 2:
        return [[], None]
    # Proyeccion
    x = np.array(x_values)
    y = np.array(y_values)
    z = np.polyfit(x, y, 1)
    p = np.poly1d(z)
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_graficos = []
    count = 0
    valor = 0
    while valor < max_aerogeneradores:
        if fecha_calculo < first_date:
            valor = 0
        else:
            valor = p(count)
            if valor < 0:
                valor = 0
            elif valor > max_aerogeneradores:
                valor = max_aerogeneradores
            count += 1
        fecha_grafico = str(fecha_calculo.year) + '-' + str(semana_calculo)
        valor_porcentaje = valor / max_aerogeneradores * 100
        data_graficos.append({"name": fecha_grafico, "y": valor_porcentaje})
        if valor < max_aerogeneradores:
            fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
            semana_calculo = fecha.isocalendar()[1]
            d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
            fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    return [data_graficos,fecha_calculo]


# En realidad es el número de componentes de un WTG para considerar que el aerogenerador ya está montado
def aerogeneradoresAMontar(parque_eolico):
    estado_montaje = EstadoFU.objects.get(nombre='Montaje')
    estado_descarga = EstadoFU.objects.get(nombre='Descarga')

    try:
        c = Componente.objects.get(nombre='Izado Cable HV')
    except Componente.DoesNotExist:
        c = None

    if c is not None:
        try:
            izado = Membership.objects.get(parque_eolico=parque_eolico,estado=estado_montaje, componente=c)
            aux = Membership.objects.filter(parque_eolico=parque_eolico, estado=estado_montaje, orden__lt=izado.orden)
            return aux.count()
        except Membership.DoesNotExist:
            logger.debug('No se encontró actividad de Izaje en parque')

    # Si no encuentro el componente Izado, entonces el número de elementos es el numero de componentes que están en
    # descarga y montaje
    aux = Membership.objects.filter(parque_eolico=parque_eolico,
                                    estado=estado_montaje)
    cuenta_total = 0
    for m in aux:
        aux2 = Membership.objects.filter(parque_eolico=parque_eolico, componente=m.componente, estado=estado_descarga)
        if aux2.count() > 0:
            cuenta_total += 1

    return cuenta_total


def aerogeneradoresMontados(parque_eolico,fecha):
    parque = parque_eolico.parque

    # aux = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
    #                                   orden_montaje=8)

    estado_montaje = EstadoFU.objects.get(nombre='Montaje')
    estado_descarga = EstadoFU.objects.get(nombre='Descarga')

    cuenta_total = aerogeneradoresAMontar(parque_eolico)
    miembros = Membership.objects.filter(parque_eolico=parque_eolico,
                                         estado=estado_montaje,
                                         orden=cuenta_total)
    try:
        componente = miembros[0].componente
    except:
        return 0


    c = Registros.objects.filter(parque=parque,
                                 componente=componente,
                                 estado=estado_montaje,
                                 fecha__lte=fecha)
    return c.count()


def porcentajeAvance(parque_eolico,fecha,componentes_montaje=None):
    parque = parque_eolico.parque
    max_aerogeneradores = parque.no_aerogeneradores
    estado_montaje = EstadoFU.objects.get(nombre='Montaje')
    miembros = None

    ag_montar = aerogeneradoresAMontar(parque_eolico)
    if ag_montar == 0:
        ag_montar = 8

    if componentes_montaje is None:
        # aux = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
        #                                   orden_montaje__range=(1, 8))

        miembros = Membership.objects.filter(parque_eolico=parque_eolico,
                                             estado=estado_montaje,
                                             orden__range=(1,ag_montar))
        componentes_montaje = []
        for m in miembros:
            componentes_montaje.append(m.componente.id)

    c = Registros.objects.filter(parque=parque,
                                 componente__in=componentes_montaje,
                                 estado=estado_montaje,
                                 fecha__lte=fecha)


    valor = float(c.count()) / ag_montar

    valor = valor / max_aerogeneradores * 100
    return valor


def graficoAvances(parque_eolico,anho,semana,actual_date,data_proyeccion):
    data_full = []
    d = str(anho) + '-W' + str(semana)
    r = datetime.strptime(d + '-0', "%Y-W%W-%w")
    s = actual_date.isocalendar()[1]
    d = str(actual_date.isocalendar()[0]) + '-W' + str(s)
    r2 = datetime.strptime(d + '-0', "%Y-W%W-%w")
    parque = parque_eolico.parque
    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return data_full
    max_aerogeneradores = parque.no_aerogeneradores
    # aux = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
    #                                                   orden_montaje__range=(1,8))
    estado_montaje = EstadoFU.objects.get(nombre='Montaje')
    aux = Membership.objects.filter(parque_eolico=parque_eolico,
                                    estado=estado_montaje,
                                    orden__range=(1, 8))
    componentes_montaje =[]
    for c in aux:
        componentes_montaje.append(c.componente.id)
    estado = EstadoFU.objects.get(idx=3)
    # Contractual
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_graficos = []
    while fecha_calculo <= r:
        c = Contractual.objects.filter(parque=parque,
                                   componente__in = componentes_montaje,
                                   estado = estado,
                                   fecha__lte= fecha_calculo).aggregate(Sum('no_aerogeneradores'))
        fecha_grafico = str(fecha_calculo.year) + '-' + str(semana_calculo)
        if c['no_aerogeneradores__sum'] is None:
            valor = 0
        else:
            valor = float(c['no_aerogeneradores__sum']) / 8 /max_aerogeneradores*100
        data_graficos.append({"name": fecha_grafico, "y": valor})
        fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
        semana_calculo = fecha.isocalendar()[1]
        d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_full.append({"name": "Contractual", "data": data_graficos})

    # Plan
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_graficos = []
    while fecha_calculo <= r:
        c = Plan.objects.filter(parque=parque,
                                       componente__in=componentes_montaje,
                                       estado=estado,
                                       fecha__lte=fecha_calculo).aggregate(Sum('no_aerogeneradores'))
        fecha_grafico = str(fecha_calculo.year) + '-' + str(semana_calculo)
        if c['no_aerogeneradores__sum'] is None:
            valor = 0
        else:
            valor = float(c['no_aerogeneradores__sum']) / 8 / max_aerogeneradores*100
        data_graficos.append({"name": fecha_grafico, "y": valor})
        fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
        semana_calculo = fecha.isocalendar()[1]
        d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_full.append({"name": "Plan", "data": data_graficos})

    # Real
    fecha = configuracion.fecha_inicio
    semana_calculo = fecha.isocalendar()[1]
    d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_graficos = []

    while fecha_calculo <= r2:
        if fecha_calculo == r2:
            valor = porcentajeAvance(parque_eolico, actual_date, componentes_montaje=componentes_montaje)
        else:
            valor = porcentajeAvance(parque_eolico,fecha_calculo,componentes_montaje=componentes_montaje)
        fecha_grafico = str(fecha_calculo.year) + '-' + str(semana_calculo)
        data_graficos.append({"name": fecha_grafico, "y": valor})
        fecha = fecha_calculo + relativedelta.relativedelta(weeks=1)
        semana_calculo = fecha.isocalendar()[1]
        d = str(fecha.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
    data_full.append({"name": "Real", "data": data_graficos})

    if len(data_proyeccion):
        data_full.append({"name": "Proyección", "data": data_proyeccion, "dashStyle": "Dot"})

    datos = serialize_grafico(data_full)
    return datos


def posicion_aerogeneradores(parque_eolico, fecha_calculo):
    parque = parque_eolico.parque
    d = str(fecha_calculo.isocalendar()[0]) + '-W' + str(fecha_calculo.isocalendar()[1])
    estado = EstadoFU.objects.get(idx=3)
    pos = OrderedDict()
    for ag in Aerogenerador.objects.filter(parque=parque):
        pos[ag.nombre] = {}

    if parque.codigo == 'CLI-001':
        pos['WTG01']['width'] = 7.0
        pos['WTG01']['top'] = 57.9
        pos['WTG01']['left'] = 6.9
        pos['WTG01']['zindex'] = 403

        pos['WTG02']['width'] = 7.0
        pos['WTG02']['top'] = 56.2
        pos['WTG02']['left'] = 12.9
        pos['WTG02']['zindex'] = 303

        pos['WTG03']['width'] = 7.0
        pos['WTG03']['top'] = 55
        pos['WTG03']['left'] = 18.5
        pos['WTG03']['zindex'] = 303

        pos['WTG04']['width'] = 7.0
        pos['WTG04']['top'] = 53.4
        pos['WTG04']['left'] = 24.0
        pos['WTG04']['zindex'] = 303

        pos['WTG05']['width'] = 7.0
        pos['WTG05']['top'] = 51.9
        pos['WTG05']['left'] = 29.5
        pos['WTG05']['zindex'] = 303

        pos['WTG06']['width'] = 7.0
        pos['WTG06']['top'] = 50.5
        pos['WTG06']['left'] = 34.9
        pos['WTG06']['zindex'] = 303

        pos['WTG07']['width'] = 7.0
        pos['WTG07']['top'] = 49.0
        pos['WTG07']['left'] = 40.1
        pos['WTG07']['zindex'] = 303

        pos['WTG08']['width'] = 7.0
        pos['WTG08']['top'] = 47.9
        pos['WTG08']['left'] = 45.5
        pos['WTG08']['zindex'] = 303
        #
        pos['WTG09']['width'] = 7.0
        pos['WTG09']['top'] = 46.5
        pos['WTG09']['left'] = 50.3
        pos['WTG09']['zindex'] = 303

        pos['WTG10']['width'] = 7.0
        pos['WTG10']['top'] = 45.4
        pos['WTG10']['left'] = 55.4
        pos['WTG10']['zindex'] = 303

        pos['WTG11']['width'] = 7.0
        pos['WTG11']['top'] = 44.3
        pos['WTG11']['left'] = 60.3
        pos['WTG11']['zindex'] = 303

        pos['WTG12']['width'] = 7.0
        pos['WTG12']['top'] = 43.0
        pos['WTG12']['left'] = 65.4
        pos['WTG12']['zindex'] = 303

        pos['WTG13']['width'] = 7.0
        pos['WTG13']['top'] = 41.8
        pos['WTG13']['left'] = 70.1
        pos['WTG13']['zindex'] = 303

        pos['WTG14']['width'] = 7.0
        pos['WTG14']['top'] = 40.6
        pos['WTG14']['left'] = 74.6
        pos['WTG14']['zindex'] = 303

        pos['WTG15']['width'] = 7.0
        pos['WTG15']['top'] = 38.9
        pos['WTG15']['left'] = 79.2
        pos['WTG15']['zindex'] = 303

        pos['WTG16']['width'] = 7.0
        pos['WTG16']['top'] = 38.0
        pos['WTG16']['left'] = 83.7
        pos['WTG16']['zindex'] = 303

        pos['WTG17']['width'] = 7.0
        pos['WTG17']['top'] = 36.6
        pos['WTG17']['left'] = 88.3
        pos['WTG17']['zindex'] = 303

        pos['WTG35']['width'] = 7.0
        pos['WTG35']['top'] = 30.6
        pos['WTG35']['left'] = 5.2
        pos['WTG35']['zindex'] = 103

        pos['WTG36']['width'] = 7.0
        pos['WTG36']['top'] = 29.5
        pos['WTG36']['left'] = 10.1
        pos['WTG36']['zindex'] = 103

        pos['WTG37']['width'] = 7.0
        pos['WTG37']['top'] = 28.5
        pos['WTG37']['left'] = 14.6
        pos['WTG37']['zindex'] = 103

        pos['WTG38']['width'] = 7.0
        pos['WTG38']['top'] = 27.4
        pos['WTG38']['left'] = 19.2
        pos['WTG38']['zindex'] = 103

        pos['WTG39']['width'] = 7.0
        pos['WTG39']['top'] = 26.3
        pos['WTG39']['left'] = 24.0
        pos['WTG39']['zindex'] = 103

        pos['WTG40']['width'] = 7.0
        pos['WTG40']['top'] = 25.5
        pos['WTG40']['left'] = 28.5
        pos['WTG40']['zindex'] = 103

        pos['WTG41']['width'] = 7.0
        pos['WTG41']['top'] = 24.8
        pos['WTG41']['left'] = 33.0
        pos['WTG41']['zindex'] = 103

        pos['WTG42']['width'] = 7.0
        pos['WTG42']['top'] = 23.5
        pos['WTG42']['left'] = 37.4
        pos['WTG42']['zindex'] = 103
        #

        pos['WTG43']['width'] = 7.0
        pos['WTG43']['top'] = 22.8
        pos['WTG43']['left'] = 41.7
        pos['WTG43']['zindex'] = 303

        pos['WTG44']['width'] = 7.0
        pos['WTG44']['top'] = 22.0
        pos['WTG44']['left'] = 45.9
        pos['WTG44']['zindex'] = 303

        pos['WTG45']['width'] = 7.0
        pos['WTG45']['top'] = 21.0
        pos['WTG45']['left'] = 50.1
        pos['WTG45']['zindex'] = 303

        pos['WTG46']['width'] = 7.0
        pos['WTG46']['top'] = 20.1
        pos['WTG46']['left'] = 54.3
        pos['WTG46']['zindex'] = 303

        pos['WTG47']['width'] = 7.0
        pos['WTG47']['top'] = 19.3
        pos['WTG47']['left'] = 58.5
        pos['WTG47']['zindex'] = 303

        pos['WTG48']['width'] = 7.0
        pos['WTG48']['top'] = 18.2
        pos['WTG48']['left'] = 62.3
        pos['WTG48']['zindex'] = 303

        pos['WTG49']['width'] = 7.0
        pos['WTG49']['top'] = 17.4
        pos['WTG49']['left'] = 66.3
        pos['WTG49']['zindex'] = 303

        pos['WTG50']['width'] = 7.0
        pos['WTG50']['top'] = 16.7
        pos['WTG50']['left'] = 70.0
        pos['WTG50']['zindex'] = 303

        pos['WTG51']['width'] = 7.0
        pos['WTG51']['top'] = 15.8
        pos['WTG51']['left'] = 74.0
        pos['WTG51']['zindex'] = 303

        pos['WTG60']['width'] = 7.0
        pos['WTG60']['top'] = 13.7
        pos['WTG60']['left'] = 38.4
        pos['WTG60']['zindex'] = 203

        pos['WTG61']['width'] = 7.0
        pos['WTG61']['top'] = 13.0
        pos['WTG61']['left'] = 42.4
        pos['WTG61']['zindex'] = 203

        pos['WTG62']['width'] = 7.0
        pos['WTG62']['top'] = 12.1
        pos['WTG62']['left'] = 46.2
        pos['WTG62']['zindex'] = 203

        pos['WTG63']['width'] = 7.0
        pos['WTG63']['top'] = 11.2
        pos['WTG63']['left'] = 50.0
        pos['WTG63']['zindex'] = 203

        pos['WTG69']['width'] = 7.0
        pos['WTG69']['top'] = 11
        pos['WTG69']['left'] = 3.9
        pos['WTG69']['zindex'] = 103

        pos['WTG70']['width'] = 7.0
        pos['WTG70']['top'] = 10.3
        pos['WTG70']['left'] = 7.9
        pos['WTG70']['zindex'] = 103

        pos['WTG71']['width'] = 7.0
        pos['WTG71']['top'] = 9.7
        pos['WTG71']['left'] = 11.9
        pos['WTG71']['zindex'] = 103

        pos['WTG72']['width'] = 7.0
        pos['WTG72']['top'] = 9.0
        pos['WTG72']['left'] = 16.0
        pos['WTG72']['zindex'] = 103

        pos['WTG73']['width'] = 7.0
        pos['WTG73']['top'] = 8.3
        pos['WTG73']['left'] = 20.0
        pos['WTG73']['zindex'] = 103

        pos['WTG74']['width'] = 7.0
        pos['WTG74']['top'] = 7.5
        pos['WTG74']['left'] = 24.0
        pos['WTG74']['zindex'] = 103

        pos['WTG75']['width'] = 7.0
        pos['WTG75']['top'] = 6.8
        pos['WTG75']['left'] = 27.8
        pos['WTG75']['zindex'] = 103

        pos['WTG76']['width'] = 7.0
        pos['WTG76']['top'] = 6.1
        pos['WTG76']['left'] = 31.6
        pos['WTG76']['zindex'] = 103

        pos['WTG77']['width'] = 7.0
        pos['WTG77']['top'] = 5.3
        pos['WTG77']['left'] = 35.2
        pos['WTG77']['zindex'] = 103

        pos['WTG78']['width'] = 7.0
        pos['WTG78']['top'] = 5
        pos['WTG78']['left'] = 38.9
        pos['WTG78']['zindex'] = 103

        pos['WTG79']['width'] = 7.0
        pos['WTG79']['top'] = 4.3
        pos['WTG79']['left'] = 42.9
        pos['WTG79']['zindex'] = 103

        pos['WTG80']['width'] = 7.0
        pos['WTG80']['top'] = 3.7
        pos['WTG80']['left'] = 46.3
        pos['WTG80']['zindex'] = 103

        pos['WTG81']['width'] = 7.0
        pos['WTG81']['top'] = 3.0
        pos['WTG81']['left'] = 50.0
        pos['WTG81']['zindex'] = 103

        pos['WTG82']['width'] = 7.0
        pos['WTG82']['top'] = 2.3
        pos['WTG82']['left'] = 53.3
        pos['WTG82']['zindex'] = 103

        pos['WTG83']['width'] = 7.0
        pos['WTG83']['top'] = 1.8
        pos['WTG83']['left'] = 56.8
        pos['WTG83']['zindex'] = 103

        pos['WTG84']['width'] = 7.0
        pos['WTG84']['top'] = 1.2
        pos['WTG84']['left'] = 60.2
        pos['WTG84']['zindex'] = 103

        pos['WTG85']['width'] = 7.0
        pos['WTG85']['top'] = 0.6
        pos['WTG85']['left'] = 63.7
        pos['WTG85']['zindex'] = 103
    elif parque.codigo == 'PCR-001': #Bicentenario
        x = 'WTG01'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.2
        pos[x]['left'] = 1.3
        pos[x]['zindex'] = 403

        x = 'WTG02'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.3
        pos[x]['left'] = 3.5
        pos[x]['zindex'] = 403

        x = 'WTG03'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 19.5
        pos[x]['left'] = 7.7
        pos[x]['zindex'] = 403

        x = 'WTG04'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 21.2
        pos[x]['left'] = 11.2
        pos[x]['zindex'] = 403

        x = 'WTG05'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.9
        pos[x]['left'] = 14.7
        pos[x]['zindex'] = 403

        x = 'WTG06'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.8
        pos[x]['left'] = 18.5
        pos[x]['zindex'] = 403

        x = 'WTG07'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 26.5
        pos[x]['left'] = 21.6
        pos[x]['zindex'] = 403

        x = 'WTG08'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 29.2
        pos[x]['left'] = 26.6
        pos[x]['zindex'] = 403

        x = 'WTG09'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 31.2
        pos[x]['left'] = 31.1
        pos[x]['zindex'] = 403

        x = 'WTG10'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 33.5
        pos[x]['left'] = 35.6
        pos[x]['zindex'] = 403

        x = 'WTG11'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 35.9
        pos[x]['left'] = 40.4
        pos[x]['zindex'] = 403

        x = 'WTG12'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 38.5
        pos[x]['left'] = 45.4
        pos[x]['zindex'] = 403

        x = 'WTG13'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 41.5
        pos[x]['left'] = 50.9
        pos[x]['zindex'] = 403

        x = 'WTG14'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 43.5
        pos[x]['left'] = 55.4
        pos[x]['zindex'] = 403

        x = 'WTG15'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 3.6
        pos[x]['left'] = 0.5
        pos[x]['zindex'] = 403

        x = 'WTG16'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 4.7
        pos[x]['left'] = 3.9
        pos[x]['zindex'] = 403

        x = 'WTG17'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 5.6
        pos[x]['left'] = 7.1
        pos[x]['zindex'] = 403

        x = 'WTG18'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 6.6
        pos[x]['left'] = 10.7
        pos[x]['zindex'] = 403

        x = 'WTG19'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 7.5
        pos[x]['left'] = 14.1
        pos[x]['zindex'] = 403

        x = 'WTG20'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 8.3
        pos[x]['left'] = 17.9
        pos[x]['zindex'] = 403

        x = 'WTG21'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 9.4
        pos[x]['left'] = 21.7
        pos[x]['zindex'] = 403

        x = 'WTG22'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 10.5
        pos[x]['left'] = 25.7
        pos[x]['zindex'] = 403

        x = 'WTG23'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 11.6
        pos[x]['left'] = 29.7
        pos[x]['zindex'] = 403

        x = 'WTG24'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 12.7
        pos[x]['left'] = 34.0
        pos[x]['zindex'] = 403

        x = 'WTG25'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 13.9
        pos[x]['left'] = 38.6
        pos[x]['zindex'] = 403

        x = 'WTG26'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 15.2
        pos[x]['left'] = 43.0
        pos[x]['zindex'] = 403

        x = 'WTG27'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.5
        pos[x]['left'] = 47.9
        pos[x]['zindex'] = 403

        x = 'WTG28'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.8
        pos[x]['left'] = 52.9
        pos[x]['zindex'] = 403

        x = 'WTG29'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 19.3
        pos[x]['left'] = 58.3
        pos[x]['zindex'] = 403

        x = 'WTG30'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.7
        pos[x]['left'] = 63.7
        pos[x]['zindex'] = 403

        x = 'WTG31'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.4
        pos[x]['left'] = 69.4
        pos[x]['zindex'] = 403

        x = 'WTG32'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.0
        pos[x]['left'] = 75.3
        pos[x]['zindex'] = 403

        x = 'WTG33'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 25.8
        pos[x]['left'] = 81.7
        pos[x]['zindex'] = 403

        x = 'WTG34'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.3
        pos[x]['left'] = 87.1
        pos[x]['zindex'] = 403
    elif parque.codigo == 'MST-002.2': # Sarco
        x = 'WTG03'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.2
        pos[x]['left'] = 43.1
        pos[x]['zindex'] = 403

        x = 'WTG05'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.2
        pos[x]['left'] = 48.3
        pos[x]['zindex'] = 403

        x = 'WTG07'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.9
        pos[x]['left'] = 53.5
        pos[x]['zindex'] = 403

        x = 'WTG08'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 30.7
        pos[x]['left'] = 42.7
        pos[x]['zindex'] = 403

        x = 'WTG09'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 29.9
        pos[x]['left'] = 46.0
        pos[x]['zindex'] = 403

        x = 'WTG11'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.9
        pos[x]['left'] = 51.8
        pos[x]['zindex'] = 403

        x = 'WTG12'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.1
        pos[x]['left'] = 54.4
        pos[x]['zindex'] = 403

        x = 'WTG13'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 26.4
        pos[x]['left'] = 56.4
        pos[x]['zindex'] = 403

        x = 'WTG15'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.3
        pos[x]['left'] = 63.4
        pos[x]['zindex'] = 403

        x = 'WTG16'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 28.7
        pos[x]['left'] = 71.1
        pos[x]['zindex'] = 403

        x = 'WTG18'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 30.7
        pos[x]['left'] = 65.4
        pos[x]['zindex'] = 403

        x = 'WTG20'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 32.3
        pos[x]['left'] = 61.1
        pos[x]['zindex'] = 403

        x = 'WTG24'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 35.6
        pos[x]['left'] = 47.9
        pos[x]['zindex'] = 403

        x = 'WTG25'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 37.7
        pos[x]['left'] = 42.5
        pos[x]['zindex'] = 403

        x = 'WTG26'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 37.0
        pos[x]['left'] = 45.3
        pos[x]['zindex'] = 403

        x = 'WTG27'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 41.8
        pos[x]['left'] = 55.5
        pos[x]['zindex'] = 403

        x = 'WTG28'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 41.3
        pos[x]['left'] = 59.1
        pos[x]['zindex'] = 403

        x = 'WTG29'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 40.4
        pos[x]['left'] = 62.2
        pos[x]['zindex'] = 403

        x = 'WTG30'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 39.4
        pos[x]['left'] = 65.0
        pos[x]['zindex'] = 403

        x = 'WTG31'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 38.3
        pos[x]['left'] = 67.8
        pos[x]['zindex'] = 403

        x = 'WTG32'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 37.5
        pos[x]['left'] = 70.7
        pos[x]['zindex'] = 403

        x = 'WTG33'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 36.2
        pos[x]['left'] = 73.2
        pos[x]['zindex'] = 403

        x = 'WTG34'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 34.3
        pos[x]['left'] = 75.5
        pos[x]['zindex'] = 403

        x = 'WTG35'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 32.7
        pos[x]['left'] = 77.4
        pos[x]['zindex'] = 403

        x = 'WTG36'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 19.3
        pos[x]['left'] = 38.2
        pos[x]['zindex'] = 403

        x = 'WTG38'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.9
        pos[x]['left'] = 42.7
        pos[x]['zindex'] = 403

        x = 'WTG39'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.6
        pos[x]['left'] = 18.5
        pos[x]['zindex'] = 403

        x = 'WTG40'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.4
        pos[x]['left'] = 16.5
        pos[x]['zindex'] = 403

        x = 'WTG41'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.1
        pos[x]['left'] = 14.8
        pos[x]['zindex'] = 403

        x = 'WTG42'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 29.9
        pos[x]['left'] = 22.4
        pos[x]['zindex'] = 403

        x = 'WTG43'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 28.7
        pos[x]['left'] = 27.0
        pos[x]['zindex'] = 403

        x = 'WTG44'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 28.0
        pos[x]['left'] = 29.3
        pos[x]['zindex'] = 403

        x = 'WTG45'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 33.6
        pos[x]['left'] = 35.0
        pos[x]['zindex'] = 403

        x = 'WTG46'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 32.7
        pos[x]['left'] = 36.7
        pos[x]['zindex'] = 403

        x = 'WTG47'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 26.6
        pos[x]['left'] = 35.0
        pos[x]['zindex'] = 403

        x = 'WTG49'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 23.6
        pos[x]['left'] = 23.4
        pos[x]['zindex'] = 403

        x = 'WTG51'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 21.8
        pos[x]['left'] = 30.4
        pos[x]['zindex'] = 403

        x = 'WTG52'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.8
        pos[x]['left'] = 33.4
        pos[x]['zindex'] = 403

        x = 'WTG54'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.0
        pos[x]['left'] = 21.3
        pos[x]['zindex'] = 403

        x = 'WTG55'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.6
        pos[x]['left'] = 23.4
        pos[x]['zindex'] = 403

        x = 'WTG57'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.2
        pos[x]['left'] = 48.9
        pos[x]['zindex'] = 403

        x = 'WTG59'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 7.2
        pos[x]['left'] = 29.6
        pos[x]['zindex'] = 403

        x = 'WTG61'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 5.0
        pos[x]['left'] = 33.1
        pos[x]['zindex'] = 403

        x = 'WTG62'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 4.7
        pos[x]['left'] = 35.1
        pos[x]['zindex'] = 403

        x = 'WTG64'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 15.8
        pos[x]['left'] = 30.8
        pos[x]['zindex'] = 403

        x = 'WTG66'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 15.2
        pos[x]['left'] = 33.3
        pos[x]['zindex'] = 403

        x = 'WTG67'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 14.6
        pos[x]['left'] = 36.0
        pos[x]['zindex'] = 403

        x = 'WTG71'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 8.0
        pos[x]['left'] = 42.1
        pos[x]['zindex'] = 403

        x = 'WTG72'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 13.0
        pos[x]['left'] = 41.4
        pos[x]['zindex'] = 403

        x = 'WTG75'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 11.8
        pos[x]['left'] = 46.4
        pos[x]['zindex'] = 403

        # x = 'WTG05'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 22.9
        # pos[x]['left'] = 14.7
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG07'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 26.5
        # pos[x]['left'] = 21.6
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG08'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 29.2
        # pos[x]['left'] = 26.6
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG09'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 31.2
        # pos[x]['left'] = 31.1
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG11'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 35.9
        # pos[x]['left'] = 40.4
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG12'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 38.5
        # pos[x]['left'] = 45.4
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG13'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 41.5
        # pos[x]['left'] = 50.9
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG15'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 3.6
        # pos[x]['left'] = 0.5
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG16'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 4.7
        # pos[x]['left'] = 3.9
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG18'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 6.6
        # pos[x]['left'] = 10.7
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG20'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 8.3
        # pos[x]['left'] = 17.9
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG24'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 12.7
        # pos[x]['left'] = 34.0
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG25'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 13.9
        # pos[x]['left'] = 38.6
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG26'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 15.2
        # pos[x]['left'] = 43.0
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG27'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 16.5
        # pos[x]['left'] = 47.9
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG28'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 17.8
        # pos[x]['left'] = 52.9
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG29'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 19.3
        # pos[x]['left'] = 58.3
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG30'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 20.7
        # pos[x]['left'] = 63.7
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG31'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 22.4
        # pos[x]['left'] = 69.4
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG32'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 24.0
        # pos[x]['left'] = 75.3
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG33'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 25.8
        # pos[x]['left'] = 81.7
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG34'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 27.3
        # pos[x]['left'] = 87.1
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG35'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 27.3
        # pos[x]['left'] = 87.1
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG36'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 27.3
        # pos[x]['left'] = 87.1
        # pos[x]['zindex'] = 403
        #
        # x = 'WTG38'
        # pos[x]['width'] = 7.0
        # pos[x]['top'] = 27.3
        # pos[x]['left'] = 87.1
        # pos[x]['zindex'] = 403
    elif parque.codigo == 'MST-002.1': # Aurora
        x = 'WTG09'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 9.0
        pos[x]['left'] = 18.7
        pos[x]['zindex'] = 403

        x = 'WTG10'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 14.2
        pos[x]['left'] = 20.4
        pos[x]['zindex'] = 403

        x = 'WTG12'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.2
        pos[x]['left'] = 7.9
        pos[x]['zindex'] = 403

        x = 'WTG13'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 21.3
        pos[x]['left'] = 12.5
        pos[x]['zindex'] = 403

        x = 'WTG14'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 12.9
        pos[x]['left'] = 29.7
        pos[x]['zindex'] = 403

        x = 'WTG15'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 12.4
        pos[x]['left'] = 32.3
        pos[x]['zindex'] = 403

        x = 'WTG16'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 11.2
        pos[x]['left'] = 35.1
        pos[x]['zindex'] = 403

        x = 'WTG17'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 10.6
        pos[x]['left'] = 37.6
        pos[x]['zindex'] = 403

        x = 'WTG18'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.3
        pos[x]['left'] = 24.9
        pos[x]['zindex'] = 403

        x = 'WTG19'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.0
        pos[x]['left'] = 27.6
        pos[x]['zindex'] = 403

        x = 'WTG20'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.1
        pos[x]['left'] = 29.6
        pos[x]['zindex'] = 403

        x = 'WTG21'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 14.9
        pos[x]['left'] = 33.8
        pos[x]['zindex'] = 403

        x = 'WTG22'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 8.5
        pos[x]['left'] = 48.3
        pos[x]['zindex'] = 403

        x = 'WTG23'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 21.6
        pos[x]['left'] = 18.7
        pos[x]['zindex'] = 403

        x = 'WTG24'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.4
        pos[x]['left'] = 23.7
        pos[x]['zindex'] = 403

        x = 'WTG26'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 25.6
        pos[x]['left'] = 19.4
        pos[x]['zindex'] = 403

        x = 'WTG27'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 19.7
        pos[x]['left'] = 33.6
        pos[x]['zindex'] = 403

        x = 'WTG28'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 25.7
        pos[x]['left'] = 23.9
        pos[x]['zindex'] = 403

        x = 'WTG29'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 23.7
        pos[x]['left'] = 28.0
        pos[x]['zindex'] = 403

        x = 'WTG30'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.0
        pos[x]['left'] = 30.8
        pos[x]['zindex'] = 403

        x = 'WTG31'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 30.6
        pos[x]['left'] = 26.9
        pos[x]['zindex'] = 403

        x = 'WTG32'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 28.6
        pos[x]['left'] = 30.3
        pos[x]['zindex'] = 403

        x = 'WTG33'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.4
        pos[x]['left'] = 40.3
        pos[x]['zindex'] = 403

        x = 'WTG34'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.3
        pos[x]['left'] = 49.0
        pos[x]['zindex'] = 403

        x = 'WTG35'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 18.1
        pos[x]['left'] = 52.3
        pos[x]['zindex'] = 403

        x = 'WTG35'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 18.1
        pos[x]['left'] = 52.3
        pos[x]['zindex'] = 403

        x = 'WTG36'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.8
        pos[x]['left'] = 54.7
        pos[x]['zindex'] = 403

        x = 'WTG37'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 12.8
        pos[x]['left'] = 61.1
        pos[x]['zindex'] = 403

        x = 'WTG38'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.9
        pos[x]['left'] = 46.3
        pos[x]['zindex'] = 403

        x = 'WTG39'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 23.7
        pos[x]['left'] = 50.1
        pos[x]['zindex'] = 403

        x = 'WTG40'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 21.7
        pos[x]['left'] = 56.3
        pos[x]['zindex'] = 403

        x = 'WTG41'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 13.6
        pos[x]['left'] = 68.3
        pos[x]['zindex'] = 403

        x = 'WTG42'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.2
        pos[x]['left'] = 54.1
        pos[x]['zindex'] = 403

        x = 'WTG43'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 13.0
        pos[x]['left'] = 70.5
        pos[x]['zindex'] = 403

        x = 'WTG44'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 15.4
        pos[x]['left'] = 74.9
        pos[x]['zindex'] = 403

        x = 'WTG45'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.4
        pos[x]['left'] = 63.2
        pos[x]['zindex'] = 403

        x = 'WTG46'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 38.4
        pos[x]['left'] = 29.7
        pos[x]['zindex'] = 403

        x = 'WTG47'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 31.2
        pos[x]['left'] = 52.6
        pos[x]['zindex'] = 403

        x = 'WTG48'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 28.9
        pos[x]['left'] = 55.1
        pos[x]['zindex'] = 403

        x = 'WTG49'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.3
        pos[x]['left'] = 62.7
        pos[x]['zindex'] = 403

        x = 'WTG50'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 25.5
        pos[x]['left'] = 65.0
        pos[x]['zindex'] = 403

        x = 'WTG51'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 39.0
        pos[x]['left'] = 35.3
        pos[x]['zindex'] = 403

        x = 'WTG52'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 31.6
        pos[x]['left'] = 61.5
        pos[x]['zindex'] = 403

        x = 'WTG53'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 46.4
        pos[x]['left'] = 44.8
        pos[x]['zindex'] = 403
    elif parque.codigo == 'EES-001.1':  # Espinal
        x = 'WTG01'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.3
        pos[x]['left'] = 5.8
        pos[x]['zindex'] = 403

        x = 'WTG02'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 18
        pos[x]['left'] = 14.8
        pos[x]['zindex'] = 403

        x = 'WTG03'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.4
        pos[x]['left'] = 19.8
        pos[x]['zindex'] = 403

        x = 'WTG04'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 15
        pos[x]['left'] = 23
        pos[x]['zindex'] = 403

        x = 'WTG05'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 14
        pos[x]['left'] = 25.5
        pos[x]['zindex'] = 403

        x = 'WTG06'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 11.5
        pos[x]['left'] = 30.6
        pos[x]['zindex'] = 403

        x = 'WTG07'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 10.5
        pos[x]['left'] = 33.6
        pos[x]['zindex'] = 403

        x = 'WTG08'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 9.4
        pos[x]['left'] = 36.9
        pos[x]['zindex'] = 403

        x = 'WTG09'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 7.7
        pos[x]['left'] = 40.2
        pos[x]['zindex'] = 403

        x = 'WTG10'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 7.2
        pos[x]['left'] = 43
        pos[x]['zindex'] = 403

        x = 'WTG11'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 6
        pos[x]['left'] = 45.5
        pos[x]['zindex'] = 403

        x = 'WTG12'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 4.3
        pos[x]['left'] = 47.9
        pos[x]['zindex'] = 403

        x = 'WTG13'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.7
        pos[x]['left'] = 11.3
        pos[x]['zindex'] = 403

        x = 'WTG14'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 26.1
        pos[x]['left'] = 14.8
        pos[x]['zindex'] = 403

        x = 'WTG15'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 20.8
        pos[x]['left'] = 26.7
        pos[x]['zindex'] = 403

        x = 'WTG16'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 18.6
        pos[x]['left'] = 31.3
        pos[x]['zindex'] = 403

        x = 'WTG17'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.5
        pos[x]['left'] = 34.1
        pos[x]['zindex'] = 403

        x = 'WTG18'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16
        pos[x]['left'] = 36.7
        pos[x]['zindex'] = 403

        x = 'WTG19'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 14.5
        pos[x]['left'] = 40.7
        pos[x]['zindex'] = 403

        x = 'WTG20'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 12.5
        pos[x]['left'] = 45.2
        pos[x]['zindex'] = 403

        x = 'WTG21'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 10.9
        pos[x]['left'] = 47.4
        pos[x]['zindex'] = 403

        x = 'WTG22'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 9.7
        pos[x]['left'] = 51.3
        pos[x]['zindex'] = 403

        x = 'WTG23'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 8.5
        pos[x]['left'] = 54.2
        pos[x]['zindex'] = 403

        x = 'WTG24'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 7.2
        pos[x]['left'] = 56.7
        pos[x]['zindex'] = 403

        x = 'WTG25'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 5.7
        pos[x]['left'] = 60.2
        pos[x]['zindex'] = 403

        x = 'WTG26'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 2.8
        pos[x]['left'] = 50
        pos[x]['zindex'] = 403

        x = 'WTG27'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 0.9
        pos[x]['left'] = 52.7
        pos[x]['zindex'] = 403

        x = 'WTG28'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 24.6
        pos[x]['left'] = 17.9
        pos[x]['zindex'] = 403

        x = 'WTG29'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 33.9
        pos[x]['left'] = 19.4
        pos[x]['zindex'] = 403

        x = 'WTG30'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 30.4
        pos[x]['left'] = 21.9
        pos[x]['zindex'] = 403

        x = 'WTG31'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 27.8
        pos[x]['left'] = 26.5
        pos[x]['zindex'] = 403

        x = 'WTG32'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 26.5
        pos[x]['left'] = 49.6
        pos[x]['zindex'] = 403

        x = 'WTG33'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 25
        pos[x]['left'] = 32.3
        pos[x]['zindex'] = 403

        x = 'WTG34'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 22.9
        pos[x]['left'] = 37.6
        pos[x]['zindex'] = 403

        x = 'WTG35'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 21.7
        pos[x]['left'] = 40.8
        pos[x]['zindex'] = 403

        x = 'WTG36'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 19.7
        pos[x]['left'] = 43.6
        pos[x]['zindex'] = 403

        x = 'WTG37'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 17.7
        pos[x]['left'] = 47.9
        pos[x]['zindex'] = 403

        x = 'WTG38'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 16.3
        pos[x]['left'] = 51.6
        pos[x]['zindex'] = 403

        x = 'WTG39'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 12
        pos[x]['left'] = 61.3
        pos[x]['zindex'] = 403

        x = 'WTG40'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 10.5
        pos[x]['left'] = 64.8
        pos[x]['zindex'] = 403

        x = 'WTG41'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 9.2
        pos[x]['left'] = 67.9
        pos[x]['zindex'] = 403

        x = 'WTG42'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 7.7
        pos[x]['left'] = 70.9
        pos[x]['zindex'] = 403

        x = 'WTG43'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 6.3
        pos[x]['left'] = 73.0
        pos[x]['zindex'] = 403

        x = 'WTG44'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 4.8
        pos[x]['left'] = 77.3
        pos[x]['zindex'] = 403

        x = 'WTG45'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 43.7
        pos[x]['left'] = 19.2
        pos[x]['zindex'] = 403

        x = 'WTG46'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 42.1
        pos[x]['left'] = 22.2
        pos[x]['zindex'] = 403

        x = 'WTG47'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 40.0
        pos[x]['left'] = 25.4
        pos[x]['zindex'] = 403

        x = 'WTG48'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 35.7
        pos[x]['left'] = 31.5
        pos[x]['zindex'] = 403

        x = 'WTG49'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 34.0
        pos[x]['left'] = 34.2
        pos[x]['zindex'] = 403

        x = 'WTG50'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 32.0
        pos[x]['left'] = 38.9
        pos[x]['zindex'] = 403

    elif parque.codigo == 'EES-001.2':  # Juchitan
        x = 'WTG01'
        pos[x]['width'] = 7.0
        pos[x]['top'] = 9.0
        pos[x]['left'] = 18.7
        pos[x]['zindex'] = 403
    else:
        return pos

    for ag in Aerogenerador.objects.filter(parque=parque,idx__gte=0):
        if len(pos[ag.nombre]) > 0:
            pos[ag.nombre]['url'] = reverse('fu:ingreso', args=(parque.slug,ag.slug))
            registros = Registros.objects.filter(parque=parque,
                                                 aerogenerador=ag,
                                                 estado=estado,
                                                 fecha__lte=fecha_calculo)
            found = False
            if registros.count() > 0:
                reg_ids = []
                for r in registros:
                    reg_ids.append(r.componente.id)
                # rel = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
                #                                   componente__in=reg_ids).order_by('-orden_montaje')
                miembros = Membership.objects.filter(parque_eolico=parque_eolico,
                                                     componente__in=reg_ids,
                                                     estado = estado).order_by('-orden')
                for m in miembros:
                    path = os.path.join(settings.BASE_DIR,'static/common/images/ag')
                    filename = path + '/' + unicode(m.componente.nombre) + '.png'
                    if os.path.isfile(unicode(filename).encode('utf-8')):
                        found = True
                        [nombre, width, top, left] = get_image_data(os.path.basename(filename),
                                                            pos[ag.nombre]['top'],
                                                            pos[ag.nombre]['left'])
                        pos[ag.nombre]['width'] = width
                        pos[ag.nombre]['top'] = top
                        pos[ag.nombre]['left'] = left
                        pos[ag.nombre]['img'] = nombre
                        break
            # Inicio Pruebas
            # path = os.path.join(settings.BASE_DIR, 'static/common/images/ag')
            # # filename = path + '/' + unicode(r.componente.nombre) + '.png'
            # filename = path + '/' + unicode('T1') + '.png'
            # # filename = path + '/' + unicode('Pala 3') + '.png'
            # if os.path.isfile(unicode(filename).encode('utf-8')):
            #     found = True
            #     [nombre, width, top, left] = get_image_data(os.path.basename(filename),
            #                                                 pos[ag.nombre]['top'],
            #                                                 pos[ag.nombre]['left'])
            #     pos[ag.nombre]['width'] = width
            #     pos[ag.nombre]['top'] = top
            #     pos[ag.nombre]['left'] = left
            #     pos[ag.nombre]['img'] = nombre
            # found = True
            # Fin Pruebas
            if not found:
                pos[ag.nombre]['width'] = 0.5
                try:
                    pos[ag.nombre]['top'] = pos[ag.nombre]['top'] + 12.7
                    pos[ag.nombre]['left'] = pos[ag.nombre]['left'] + 3.7
                except:
                    pass;

                pos[ag.nombre]['img'] = 'common/images/ag/0.png'

    return pos


def get_image_data(filename, top_i, left_i):
    name = '0.png'
    try:
        top = top_i + 12.7
    except:
        pass;
    left = left_i + 3.7
    width = 0.5

    if filename == 'Pala 3.png':
        name = filename
        width = 4.5
        top = top_i + 5.4
        left = left_i + 1.4
    elif filename == 'T1.png':
        name = filename
        width = 0.5
        top = top_i + 12.7
        left = left_i + 3.7
    elif filename == 'T2.png':
        name = filename
        width = 0.6
        top = top_i + 11.1
        left = left_i + 3.7
    elif filename == 'T3.png':
        name = filename
        width = 0.5
        top = top_i + 10.1
        left = left_i + 3.7
    elif filename == 'Nacelle.png':
        name = filename
        width = 0.65
        top = top_i + 8.8
        left = left_i + 3.6
    elif filename == 'Buje.png':
        name = filename
        width = 0.8
        top = top_i + 8.7
        left = left_i + 3.7
    elif filename == 'Pala 1.png':
        name = filename
        width = 1.1
        top = top_i + 5.8
        left = left_i + 3.4
    elif filename == 'Pala 2.png':
        name = filename
        width = 4.2
        top = top_i + 8.4
        left = left_i + 1.4
    elif filename == 'Rotor.png':
        name = filename
        width = 4.5
        top = top_i + 5.4
        left = left_i + 1.4
    return ['common/images/ag/'+name, width,top,left]


def get_plano3d_img(parque):
    if parque.codigo == 'PCR-001':
        return  ['/common/images/plano-bicentenario.png', '0px', '85%', '0px']
    elif parque.codigo == 'MST-002.2':
        return ['/common/images/plano-sarco2.png', '50px', '2%', '80px']
    elif parque.codigo == 'MST-002.1':
        return ['/common/images/plano-aurora.png', '50px', '85%', '80px']
    elif parque.codigo == 'EES-001.1':
        return ['/common/images/plano-espinal.jpg', '50px', '85%', '80px']
    elif parque.codigo == 'EES-001.2':
        return ['/common/images/plano_juchitan.jpg', '50px', '85%', '80px']
    else:
        return ['/common/images/plano2.png', '0px', '85%', '0px']


@login_required(login_url='ingresar')
def dashboard(request, slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    contenido = ContenidoContainer()
    contenido.user = request.user
    contenido.titulo = u'Dashboard Follow Up'
    contenido.subtitulo = u'Parque Eólico - ' + parque.nombre
    contenido.menu = ['menu-fu', 'menu2-dashboard']

    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return TemplateResponse(request, 'fu/dashboard.html',
                      {'cont': contenido,
                       'parque': parque,
                       'aerogeneradores': aerogeneradores,
                       'configuracion': None,
                       })

    t = datetime.now()
    anho = t.isocalendar()[0]
    semana = t.isocalendar()[1]
    semana_today = t.isocalendar()[1]
    if request.method == 'POST':
        if 'semana' in request.POST:
            values = request.POST['semana'].split('-')
            anho = int(values[0])
            semana = int(values[1])
    d = str(anho) + '-W' + str(semana)
    last_day_week = datetime.strptime(d + '-0', "%Y-W%W-%w")

    estado = EstadoFU.objects.get(idx=1)
    graficoDescarga = graficoComponentes(parque_eolico,estado,last_day_week)
    estado = EstadoFU.objects.get(idx=3)
    graficoMontaje = graficoComponentes(parque_eolico, estado, last_day_week)
    estado = EstadoFU.objects.get(idx=4)
    graficoPuestaenMarcha = graficoComponentes(parque_eolico, estado, last_day_week)
    [proyeccion, last_week] = calcularProyeccionGrafico(parque_eolico, anho, semana)
    if last_week is not None:
        graficoAvance = graficoAvances(parque_eolico, last_week.isocalendar()[0], last_week.isocalendar()[1], last_day_week, proyeccion)
    else:
        fecha_aux = configuracion.fecha_final
        semana_calculo = fecha_aux.isocalendar()[1]
        d = str(fecha_aux.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
        graficoAvance = graficoAvances(parque_eolico, fecha_calculo.isocalendar()[0], fecha_calculo.isocalendar()[1], last_day_week,
                                       proyeccion)

    thisweek = str(anho) + "-" + str(semana)
    week_str = 'Semana ' + str(semana)
    if semana == semana_today:
        fecha = t
    else:
        d = str(anho) + '-W' + str(semana)
        fecha = datetime.strptime(d + '-0', "%Y-W%W-%w")
    pos_ag = posicion_aerogeneradores(parque_eolico, last_day_week)

    fecha_aux= configuracion.fecha_inicio #+ relativedelta.relativedelta(weeks=1)
    semana_calculo = fecha_aux.isocalendar()[1]
    d = str(fecha_aux.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_inicial = datetime.strptime(d + '-0', "%Y-W%W-%w")

    avance = porcentajeAvance(parque_eolico,last_day_week)
    montados = aerogeneradoresMontados(parque_eolico, last_day_week)
    try:
        componente = Componente.objects.get(nombre="Mechanical Completion")
        estado = EstadoFU.objects.get(idx=3)
        c = Registros.objects.filter(parque=parque,
                                     componente=componente,
                                     estado=estado,
                                     fecha__lte=last_day_week)
        mechanical = c.count()
    except Componente.DoesNotExist:
        mechanical = 0

    try:
        componente = Componente.objects.get(nombre="Commisioning")
        estado = EstadoFU.objects.get(idx=4)
        c = Registros.objects.filter(parque=parque,
                                     componente=componente,
                                     estado=estado,
                                     fecha__lte=last_day_week)
        ag_ready = c.count()
    except Componente.DoesNotExist:
        ag_ready = 0

    graficoAvancePost = '{series:' + graficoAvance + '}'
    logger.debug(graficoAvancePost)
    logger.debug(thisweek)

    [plano_3d, plano_3d_top, boton_left, plano_3d_top_zoom] = get_plano3d_img(parque)

    return TemplateResponse(request, 'fu/dashboard.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'graficoDescarga': graficoDescarga,
                   'graficoMontaje': graficoMontaje,
                   'graficoPuestaenMarcha':graficoPuestaenMarcha,
                   'graficoAvance': graficoAvance,
                   'thisweek': thisweek,
                   'week_str': week_str,
                   'fecha': fecha,
                   'pos_ag': pos_ag,
                   'configuracion': configuracion,
                   'fecha_inicial' : fecha_inicial,
                   'avance':avance,
                   'montados':montados,
                   'mechanical': mechanical,
                   'ag_ready': ag_ready,
                   'plano_3d': plano_3d,
                   'plano_3d_top': plano_3d_top,
                   'boton_left': boton_left,
                   'plano_3d_top_zoom': plano_3d_top_zoom,
                   })


@login_required(login_url='ingresar')
def dashboard_imagen(request, slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    contenido = ContenidoContainer()
    contenido.user = request.user
    contenido.titulo = u'Dashboard Follow Up'
    contenido.subtitulo = u'Parque Eólico - ' + parque.nombre
    contenido.menu = ['menu-fu', 'menu2-dashboard']

    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return TemplateResponse(request, 'fu/dashboard.html',
                      {'cont': contenido,
                       'parque': parque,
                       'aerogeneradores': aerogeneradores,
                       'configuracion': None,
                       })

    [plano_3d, plano_3d_top, boton_left, plano_3d_top_zoom] = get_plano3d_img(parque)

    return TemplateResponse(request, 'fu/dashboard-imagen.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'plano_3d': plano_3d,
                   'plano_3d_top': plano_3d_top,
                   'boton_left': boton_left,
                   'plano_3d_top_zoom': plano_3d_top_zoom,
                   })


@login_required(login_url='ingresar')
def dashboard_diario(request, slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    contenido = ContenidoContainer()
    contenido.user = request.user
    contenido.titulo = u'Dashboard Follow Up'
    contenido.subtitulo = u'Parque Eólico - ' + parque.nombre
    contenido.menu = ['menu-fu', 'menu2-dashboard-diario']

    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return TemplateResponse(request, 'fu/dashboard-diario.html',
                      {'cont': contenido,
                       'parque': parque,
                       'aerogeneradores': aerogeneradores,
                       'configuracion': None,
                       })

    t = datetime.now()
    semana_today = t.isocalendar()[1]
    if request.method == 'POST':
        if 'fecha' in request.POST:
            t = datetime.strptime(request.POST['fecha'],"%d-%m-%Y")
    anho = t.isocalendar()[0]
    semana = t.isocalendar()[1]
    last_day_week = t

    estado = EstadoFU.objects.get(idx=1)
    graficoDescarga = graficoComponentes(parque_eolico, estado, t)
    estado = EstadoFU.objects.get(idx=3)
    graficoMontaje = graficoComponentes(parque_eolico, estado, t)
    estado = EstadoFU.objects.get(idx=4)
    graficoPuestaenMarcha = graficoComponentes(parque_eolico, estado, t)
    [proyeccion, last_week] = calcularProyeccionGrafico(parque_eolico, anho, semana)
    if last_week is not None:
        graficoAvance = graficoAvances(parque_eolico,
                                       last_week.isocalendar()[0],
                                       last_week.isocalendar()[1],
                                       t,
                                       proyeccion)
    else:
        fecha_aux = configuracion.fecha_final
        semana_calculo = fecha_aux.isocalendar()[1]
        d = str(fecha_aux.isocalendar()[0]) + '-W' + str(semana_calculo)
        fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")
        graficoAvance = graficoAvances(parque_eolico,
                                       fecha_calculo.isocalendar()[0],
                                       fecha_calculo.isocalendar()[1],
                                       t,
                                       proyeccion)

    thisweek = str(anho) + "-" + str(semana)
    week_str = 'Semana ' + str(semana)
    fecha = t
    pos_ag = posicion_aerogeneradores(parque_eolico,t)

    fecha_aux= configuracion.fecha_inicio #+ relativedelta.relativedelta(weeks=1)
    semana_calculo = fecha_aux.isocalendar()[1]
    d = str(fecha_aux.isocalendar()[0]) + '-W' + str(semana_calculo)
    fecha_inicial = datetime.strptime(d + '-0', "%Y-W%W-%w")

    avance = porcentajeAvance(parque_eolico,last_day_week)
    montados = aerogeneradoresMontados(parque_eolico,last_day_week)
    try:
        componente = Componente.objects.get(nombre="Mechanical Completion")
        estado = EstadoFU.objects.get(idx=3)
        c = Registros.objects.filter(parque=parque,
                                     componente=componente,
                                     estado=estado,
                                     fecha__lte=last_day_week)
        mechanical = c.count()
    except Componente.DoesNotExist:
        mechanical = 0

    try:
        componente = Componente.objects.get(nombre="Commisioning")
        estado = EstadoFU.objects.get(idx=4)
        c = Registros.objects.filter(parque=parque,
                                     componente=componente,
                                     estado=estado,
                                     fecha__lte=last_day_week)
        ag_ready = c.count()
    except Componente.DoesNotExist:
        ag_ready = 0

    [plano_3d, plano_3d_top, boton_left, plano_3d_top_zoom] = get_plano3d_img(parque)

    return TemplateResponse(request, 'fu/dashboard-diario.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'graficoDescarga': graficoDescarga,
                   'graficoMontaje': graficoMontaje,
                   'graficoPuestaenMarcha':graficoPuestaenMarcha,
                   'graficoAvance': graficoAvance,
                   'thisweek': thisweek,
                   'week_str': week_str,
                   'fecha': fecha,
                   'pos_ag': pos_ag,
                   'configuracion': configuracion,
                   'fecha_inicial' : fecha_inicial,
                   'avance':avance,
                   'montados':montados,
                   'mechanical': mechanical,
                   'ag_ready': ag_ready,
                   'plano_3d': plano_3d,
                   'plano_3d_top': plano_3d_top,
                   'boton_left': boton_left,
                   'plano_3d_top_zoom': plano_3d_top_zoom,
                   })


@login_required(login_url='ingresar')
def avance(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    contenido = ContenidoContainer()
    contenido.user = request.user
    contenido.titulo = u'Avance'
    contenido.subtitulo = u'Parque Eólico - ' + parque.nombre
    contenido.menu = ['menu-fu', 'menu2-avance']

    pos_ag = posicion_aerogeneradores(parque_eolico, 2017, 39)

    return TemplateResponse(request, 'fu/avance.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'pos_ag': pos_ag,
                   })


@login_required(login_url='ingresar')
@permission_required('fu.add_componente', raise_exception=True)
def componente(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Componentes Follow Up'
    contenido.subtitulo=u'Parque Eólico - ' + parque.nombre
    contenido.menu = ['menu-fu', 'menu2-componentes']

    if request.method == 'POST':
        if 'delete' in request.POST:
            logger.debug('Componente a eliminar id=' + request.POST['del_id'])
            c = get_object_or_404(Componente, id=int(request.POST['del_id']))
            log_msg = "Se elimina componente para parque " + parque.nombre + \
                      " - Nombre " + c.nombre
            log = Log(texto=log_msg, tipo=3, user=request.user)
            log.save()
            c.delete()
            messages.add_message(request, messages.SUCCESS, 'Componente eliminado.')
        elif 'id' in request.POST:
            logger.debug('Componente a editar id=' + request.POST['id'])
            c = get_object_or_404(Componente, id=int(request.POST['id']))
            inputForm = ComponenteForm(request.POST)
            if inputForm.is_valid():
                c.nombre = inputForm.cleaned_data['nombre']
                c.estados.clear()
                for e in inputForm.cleaned_data['estadofu']:
                    c.estados.add(e)
                c.save()
                log_msg = "Se edita componente para parque " + parque.nombre + \
                          " - Nombre " + c.nombre
                log = Log(texto=log_msg, tipo=2, user=request.user)
                log.save()
        else:
            inputForm = ComponenteForm(request.POST)
            if inputForm.is_valid():
                logger.debug('Formulario válido')
                c = Componente(nombre=inputForm.cleaned_data['nombre'])
                c.save()
                for e in inputForm.cleaned_data['estadofu']:
                    c.estados.add(e)
                c.save()
                log_msg = "Se agrega componente para parque " + parque.nombre + \
                          " - Nombre " + c.nombre
                log = Log(texto=log_msg, tipo=1, user=request.user)
                log.save()
                messages.add_message(request, messages.SUCCESS, 'Componente Creado con éxito!')
            else:
                messages.add_message(request, messages.ERROR, 'Componente no pudo ser creado.')

    componenteForm = ComponenteForm()
    componentes = Componente.objects.all().order_by('id')
    return TemplateResponse(request, 'fu/componentes.html',
        {'cont': contenido,
         'parque': parque,
         'aerogeneradores':aerogeneradores,
         'componenteForm': componenteForm,
         'componentes': componentes,
        })


def getOrden(componente, d,p,m,pm):
    ret_d = 0
    ret_p = 0
    ret_m = 0
    ret_pm = 0
    for e in componente.estados.all():
        if e.idx == 1:
            ret_d = d + 1
        elif e.idx == 2:
            ret_p = p + 1
        elif e.idx == 3:
            ret_m = m + 1
        elif e.idx == 4:
            ret_pm = pm + 1
    return [ret_d,ret_p,ret_m,ret_pm]


def getComponentesbyState(parque_eolico, estado):
    if estado == 'descarga':
        e = EstadoFU.objects.get(nombre='Descarga')
        orden = 'orden_descarga'
        indice = 1
    elif estado == 'premontaje':
        e = EstadoFU.objects.get(nombre='Pre-montaje')
        orden = 'orden_premontaje'
        indice = 2
    elif estado == 'montaje':
        e = EstadoFU.objects.get(nombre='Montaje')
        orden = 'orden_montaje'
        indice = 3
    elif estado == 'puestaenmarcha':
        e = EstadoFU.objects.get(nombre='Puesta en marcha')
        orden = 'orden_puestaenmarcha'
        indice = 4
    else:
        return None

    lista = OrderedDict()

    members = parque_eolico.membership_set.filter(estado = e).order_by('orden')
    for m in members:
        lista[m.componente.id] = m.componente.nombre
    return lista
    # componentes = componentes_parque.componentes.all()
    # filtro = 'relacionesfu__' + orden
    # karws = {filtro + '__gt': 0}
    # id = 0
    # for c in componentes.filter(**karws).order_by(filtro):
    #     lista[c.id]=c.nombre
    #     id += 1
    # return lista


def addComponente(parque_eolico, new_componente, estados):
    #print(new_componente)
    for e in estados:
        aux = Membership(parque_eolico=parque_eolico,
                         componente=new_componente,
                         estado = e)
        aux.save()
    #aux = RelacionesFU.objects.filter(componentes_parque=componentes_parque)
    # if aux.exists():
    #     maximos = aux.aggregate(Max('orden_descarga'),
    #                             Max('orden_premontaje'),
    #                             Max('orden_montaje'),
    #                             Max('orden_puestaenmarcha'))
    #     descarga = maximos['orden_descarga__max']
    #     premontaje = maximos['orden_premontaje__max']
    #     montaje = maximos['orden_montaje__max']
    #     puestaenmarcha = maximos['orden_puestaenmarcha__max']
    # else:
    #     descarga = 0
    #     premontaje = 0
    #     montaje = 0
    #     puestaenmarcha = 0
    # [descarga, premontaje, montaje, puestaenmarcha] = getOrden(new_componente,
    #                                                            descarga,
    #                                                            premontaje,
    #                                                            montaje,
    #                                                            puestaenmarcha)
    #
    # r = RelacionesFU(componentes_parque=componentes_parque,
    #                  componente=new_componente,
    #                  orden_descarga=descarga,
    #                  orden_premontaje=premontaje,
    #                  orden_montaje=montaje,
    #                  orden_puestaenmarcha=puestaenmarcha)
    # r.save()


def deleteComponente(parque_eolico, del_componente):
    elementos = Membership.objects.filter(parque_eolico=parque_eolico, componente=del_componente)
    for m in elementos:
        m.delete()
    fixOrdenComponentes(parque_eolico)
    # aux = RelacionesFU.objects.get(componentes_parque=componentes_parque, componente=del_componente)
    # if aux.orden_descarga > 0:
    #     aux2=RelacionesFU.objects.filter(componentes_parque=componentes_parque, orden_descarga__gt = aux.orden_descarga)
    #     for c in aux2:
    #         c.orden_descarga += -1
    #         c.save()
    # if aux.orden_premontaje > 0:
    #     aux2=RelacionesFU.objects.filter(componentes_parque=componentes_parque, orden_premontaje__gt = aux.orden_premontaje)
    #     for c in aux2:
    #         c.orden_premontaje += -1
    #         c.save()
    # if aux.orden_montaje > 0:
    #     aux2=RelacionesFU.objects.filter(componentes_parque=componentes_parque, orden_montaje__gt = aux.orden_montaje)
    #     for c in aux2:
    #         c.orden_montaje += -1
    #         c.save()
    # if aux.orden_puestaenmarcha > 0:
    #     aux2=RelacionesFU.objects.filter(componentes_parque=componentes_parque, orden_puestaenmarcha__gt = aux.orden_puestaenmarcha)
    #     for c in aux2:
    #         c.orden_puestaenmarcha += -1
    #         c.save()
    # aux.delete()
    # fixOrdenComponentes()


def fixOrdenComponentes(parque_eolico):
    estado = EstadoFU.objects.all()
    for e in estado:
        miembros = Membership.objects.filter(parque_eolico=parque_eolico, estado=e).order_by('orden')
        orden = 1
        for m in miembros:
            m.orden = orden
            m.save()
            orden += 1
    # relaciones = ComponentesParque.objects.all()
    # filter_values = ['orden_descarga','orden_premontaje','orden_montaje','orden_puestaenmarcha']
    # for relacion in relaciones:
    #     for f in filter_values:
    #         karws = {f + '__gt': 0}
    #         componentes = relacion.relacionesfu_set.filter(**karws).order_by(f)
    #         pos = 1
    #         for c in componentes:
    #             setattr(c, f, pos)
    #             c.save()
    #             pos += 1


@login_required(login_url='ingresar')
@permission_required('fu.add_componentesparque', raise_exception=True)
def actividades(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Actividades Follow Up'
    contenido.subtitulo=u'Parque Eólico - ' + parque.nombre
    contenido.menu = ['menu-fu', 'menu2-actividades']

    if request.method == 'POST':
        if 'delComponente' in request.POST:
            deleteComponentesForm = DeleteComponentesForm(request.POST, parque=parque)
            if deleteComponentesForm.is_valid():
                log_msg = "Se elimina componente del proyecto (actividad) para parque " + parque.nombre + \
                          " - Nombre " + deleteComponentesForm.cleaned_data['componente'].nombre

                deleteComponente(parque_eolico,deleteComponentesForm.cleaned_data['componente'])

                log = Log(texto=log_msg, tipo=3, user=request.user)
                log.save()
        elif 'addComponente' in request.POST:
            choicesComponentesForm = AddComponentesForm(request.POST, parque=parque)
            if choicesComponentesForm.is_valid():
                log_msg = "Se agrega componente al proyecto (actividad) para parque " + parque.nombre + \
                          " - Nombre " + choicesComponentesForm.cleaned_data['componente'].nombre

                addComponente(parque_eolico,
                              choicesComponentesForm.cleaned_data['componente'],
                              choicesComponentesForm.cleaned_data['estadofu'])
                log = Log(texto=log_msg, tipo=1, user=request.user)
                log.save()

    choicesComponentesForm = AddComponentesForm(parque = parque)
    deleteComponentesForm = DeleteComponentesForm(parque = parque)
    lista_descarga = getComponentesbyState(parque_eolico,'descarga')
    lista_premontaje = getComponentesbyState(parque_eolico,'premontaje')
    lista_montaje = getComponentesbyState(parque_eolico,'montaje')
    lista_puestaenmarcha = getComponentesbyState(parque_eolico,'puestaenmarcha')

    actividades = OrderedDict()
    actividades['descarga']=lista_descarga
    actividades['premontaje']= lista_premontaje
    actividades['montaje']= lista_montaje
    actividades['puestaenmarcha']= lista_puestaenmarcha

    titulos = {
        'descarga': 'Descarga',
        'premontaje': 'Pre-montaje',
        'montaje': 'Montaje',
        'puestaenmarcha': 'Puesta en marcha'
    }

    return TemplateResponse(request, 'fu/actividades.html',
        {'cont': contenido,
         'parque': parque,
         'aerogeneradores':aerogeneradores,
         'choicesComponentesForm': choicesComponentesForm,
         'deleteComponentesForm': deleteComponentesForm,
         'actividades': actividades,
         'titulos': titulos,
        })


@csrf_exempt
def status_componentes(request, slug):
    if request.is_ajax() and request.POST:
        componente_id = request.POST.get('id')
        datos = []
        #componente = Componente.objects.get(id=componente_id)
        for e in Componente.objects.get(id=componente_id).estados.all():
            datos.append(e.id)
        return HttpResponse(json.dumps(datos), content_type='application/json')


@csrf_exempt
def ordenar_actividades(request, slug, str_estado):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    parque_eolico = get_object_or_404(ParqueEolico, parque=parque)
    logger.debug("Ordenar Actividades, estado = " + str_estado)
    response_data = {}
    response_data['files'] = []
    if str_estado == 'descarga':
        estado = EstadoFU.objects.get(nombre='Descarga')
    elif str_estado == 'premontaje':
        estado = EstadoFU.objects.get(nombre='Pre-Montaje')
    elif str_estado == 'montaje':
        estado = EstadoFU.objects.get(nombre='Montaje')
    elif str_estado == 'puestaenmarcha':
        estado = EstadoFU.objects.get(nombre='Puesta en marcha')

    if request.method == 'POST':
        response = json.dumps([])
        post_dict = parser.parse(request.POST.urlencode())
        datos = post_dict['lista']
        for pos, val in datos.iteritems():
            idx = pos + 1
            id = int(val['id'])
            componente = Componente.objects.get(id=id)
            obj = Membership.objects.get(parque_eolico=parque_eolico, componente=componente, estado=estado)
            obj.orden = idx
            obj.save()
        log_msg = "Cambio en orden de componentes para parque " + parque.nombre
        log = Log(texto=log_msg, tipo=2, user=request.user)
        log.save()
        return HttpResponse(
            response,
            content_type="application/json"
        )


def checkValidFile(ws, parque_eolico):
    fila = 4
    columna = 2
    valor = ws.cell(row=fila,column=columna).value
    while valor is not None:
        elementos = parque_eolico.componentes.filter(nombre=valor)
        if elementos.count() <= 0:
            logger.debug('Archivo inválido')
            return -1
        fila = fila +2
        valor = ws.cell(row=fila, column=columna).value

    return fila


def getLastColumn(ws):
    fila = 3
    columna = 4
    valor = ws.cell(row=fila,column=columna).value
    while valor is not None:
        # Verifica que los nombres sean componentes válidos
        columna = columna + 2
        valor = ws.cell(row=fila, column=columna).value
    return columna


@login_required(login_url='ingresar')
@permission_required('fu.add_configuracionfu', raise_exception=True)
def configuracion(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        configuracion = None

    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Parque Eólico'
    contenido.subtitulo= parque.nombre
    contenido.menu = ['menu-fu', 'menu2-configuracionfu']

    form = None

    if request.method == 'POST':
        if configuracion is None:
            form = ConfiguracionFUForm(request.POST)
        else:
            form = ConfiguracionFUForm(request.POST, instance=configuracion)
        if form.is_valid():
            if configuracion is None:
                conf = form.save(commit=False)
                conf.parque = parque
                conf.save()
                log_msg = "Se cambia configuración Follow Up para parque " + parque.nombre
                log = Log(texto=log_msg, tipo=2, user=request.user)
                log.save()
            else:
                form.save()
            messages.add_message(request, messages.SUCCESS, 'Configuración guardada con éxito')
        else:
            messages.add_message(request, messages.ERROR, 'Configuración no se ha podido guardar.')

    if form is None:
        if configuracion is None:
            form = ConfiguracionFUForm()
        else:
            form = ConfiguracionFUForm(instance=configuracion)
    return TemplateResponse(request, 'fu/configuracion.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'form': form
                   })


def readPlanFile(configuracion, parque_eolico):
    nombre_archivo = os.path.join(settings.MEDIA_ROOT, configuracion.plan.name)
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    last_row = checkValidFile(ws, parque_eolico)
    if last_row == -1:
        # Retorna falso, no se pudo leer este archivo correctamente.
        return False
    else:
        # si está bien el archivo, entonces hay que borrar la planificación para el parque antes de insertar la nueva.
        objs = Plan.objects.filter(parque = configuracion.parque)
        for obj in objs:
            obj.delete()
        objs = Contractual.objects.filter(parque = configuracion.parque)
        for obj in objs:
            obj.delete()
    last_column = getLastColumn(ws)
    fila = 4
    ultimo_estado = None
    ultimo_anho = None
    estado = None
    while fila < last_row:
        columna = 4
        nombre_componente = ws.cell(row=fila, column=2).value
        # lo puedo hacer de forma segura, porque se supone que los validé en checkValidFile
        componente = Componente.objects.get(nombre=nombre_componente)
        nombre_estado = ws.cell(row=fila, column=1).value

        if nombre_estado is not None:
            if nombre_estado != ultimo_estado:
                ultimo_estado = nombre_estado
                estado = None
                if ultimo_estado == 'Descarga en Parque':
                    estado = EstadoFU.objects.get(idx=1)
                elif ultimo_estado == 'Pre-montaje':
                    estado = EstadoFU.objects.get(idx=2)
                elif ultimo_estado == 'Montaje':
                    estado = EstadoFU.objects.get(idx=3)
                elif ultimo_estado == 'Puesta en marcha':
                    estado = EstadoFU.objects.get(idx=4)
                else:
                    estado = None
                    logger.debug('No existe el estado')

        if ultimo_estado is not None and estado is not None:
            while columna < last_column:
                semana = ws.cell(row=3, column=columna).value
                anho = ws.cell(row=1, column=columna).value
                if anho is not None:
                    if anho != ultimo_anho:
                        ultimo_anho = anho
                contractual = ws.cell(row=fila,column = columna).value
                plan = ws.cell(row=fila+1,column = columna).value
                d = ultimo_anho + "-W" + semana + "-0"
                fecha = datetime.strptime(d, "%Y-W%W-%w")
                if plan is not None:
                    plan = int(plan)
                    if plan != 0:
                        nuevo = Plan(parque = configuracion.parque,
                                     componente = componente,
                                     estado = estado,
                                     fecha = fecha,
                                     no_aerogeneradores=plan)
                        nuevo.save()

                if contractual is not None:
                    contractual = int(contractual)
                    if contractual != 0:
                        nuevo = Contractual(parque = configuracion.parque,
                                     componente = componente,
                                     estado= estado,
                                     fecha = fecha,
                                     no_aerogeneradores=contractual)
                        nuevo.save()

                columna += 1
        fila += 2
    return True


def graficoPlanificacion(parque):
    configuracion = ConfiguracionFU.objects.get(parque=parque)
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    aux = configuracion.fecha_inicio
    final = configuracion.fecha_final
    xvalues = {}
    yvalues = {}
    plan_values = {}
    contractual_values = {}
    for i in range(1,5):
        xlabels = '['
        ylabels = '['
        plan = '['
        contractual = '['
        idx = 0
        idy = 0
        first = True
        aux = configuracion.fecha_inicio
        while aux < final:
            semana = str(aux.isocalendar()[1])
            anho = aux.isocalendar()[0]
            d = str(anho) + "-W" + semana + "-0"
            fecha_query = datetime.strptime(d, "%Y-W%W-%w")
            xlabels += '"'+ str(anho) + '-' + semana + '",'
            idy = 0
            for componente in parque_eolico.componentes.all().distinct():
                e = None
                for aux_e in componente.estados.all():
                    if aux_e.idx == i:
                        e = aux_e
                        break
                if e is not None:
                    if first:
                        ylabels += '"' + componente.nombre + '",'
                    try:
                        q = Plan.objects.get(parque=parque, componente=componente, estado=e,fecha=fecha_query)
                        val = q.no_aerogeneradores
                    except Plan.DoesNotExist:
                        val = 0
                    plan += '['+str(idx)+','+str(idy)+','+str(val)+'],'
                    try:
                        q = Contractual.objects.get(parque=parque, componente=componente, estado=e,fecha=fecha_query)
                        val = q.no_aerogeneradores
                    except Contractual.DoesNotExist:
                        val = 0
                    contractual += '['+str(idx)+','+str(idy)+','+str(val)+'],'
                    idy += 1
            first = False
            idx += 1
            aux = aux + relativedelta.relativedelta(weeks=1)
        xlabels = xlabels[:-1] + ']'
        ylabels = ylabels[:-1] + ']'
        plan = plan[:-1] + ']'
        contractual = contractual[:-1] + ']'
        xvalues[i] = xlabels
        yvalues[i] = ylabels
        plan_values[i] = plan
        contractual_values[i] = contractual
    return [xvalues,yvalues,plan_values,contractual_values]


@login_required(login_url='ingresar')
def planificacion(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo= u'Planificación Follow Up'
    contenido.subtitulo= u'Parque Eólico ' +parque.nombre
    contenido.menu = ['menu-fu', 'menu2-planificacion']

    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        configuracion = None


    form = None

    if request.method == 'POST':
        if not request.user.has_perm('fu.add_plan'):
            raise PermissionDenied
        form = PlanificacionForm(request.POST, request.FILES, instance=configuracion)
        if form.is_valid():
            configuracion = form.save()

        if configuracion.plan:
            if configuracion.plan != configuracion.prev_plan:
                readPlanFile(configuracion, parque_eolico)
                log_msg = "Se agrega planificación para parque " + parque.nombre
                log = Log(texto=log_msg, tipo=1, user=request.user)
                log.save()
            else:
                logger.debug('Se sube configuración, pero se mantiene igual.')

    if form is None and configuracion is not None:
        form = PlanificacionForm(instance = configuracion)

    if configuracion:
        [x_axis, y_axis, plan,contractual] = graficoPlanificacion(parque)
    else:
        x_axis = None
        y_axis = None
        plan = None
        contractual = None

    aux = datetime.today()
    semana = str(aux.isocalendar()[1])
    anho = aux.isocalendar()[0]
    thisweek = str(anho) + "-" + semana

    return TemplateResponse(request, 'fu/planificacion.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'configuracion': configuracion,
                   'form': form,
                   'x_axis': x_axis,
                   'y_axis': y_axis,
                   'plan': plan,
                   'contractual': contractual,
                   'thisweek': thisweek
                   })


@login_required(login_url='ingresar')
def download_config(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        return

    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    aux = configuracion.fecha_inicio
    final = configuracion.fecha_final

    semanas = []
    meses = {}
    anhos = {}
    semana = str(aux.isocalendar()[1])
    semanas.append(semana)
    meses[semana] = meses_espanol[str(aux.month)]
    anhos[semana] = str(aux.isocalendar()[0])
    aux = aux + relativedelta.relativedelta(weeks=1)
    while aux < final:
        semana = str(aux.isocalendar()[1])
        semanas.append(semana)
        meses[semana] = meses_espanol[str(aux.month)]
        anhos[semana] = str(aux.isocalendar()[0])
        aux = aux + relativedelta.relativedelta(weeks=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Planificacion"
    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    thin = Side(border_style="thin", color="000000")
    solid = Side(border_style="medium", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)
    bordertop = Border(top=thin, left=thin, right=thin)
    borderbottom = Border(left=thin, right=thin, bottom=thin)
    borderside = Border(left=thin, right=thin)
    # alignment = Alignment(horizontal="center", vertical="center",text_rotation =90, wrap_text = False, shrink_to_fit = False, indent = 0)
    alignment1 = Alignment(horizontal="center", vertical="center", text_rotation=90)
    alignment2 = Alignment(vertical="center")
    alignment3 = Alignment(horizontal="center")
    alignment4 = Alignment(horizontal="center", wrap_text = True)

    row = 3
    column = 4
    idx = 0

    d = ws.cell(row=row, column=1, value='Actividad')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=2, value='Componente')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=3, value='Semana')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    ws.column_dimensions["A"].width = 11.0
    ws.column_dimensions["B"].width = 25.0
    ws.column_dimensions["C"].width = 11.0
    prev_year = ''
    prev_month = ''
    prev_month_col = 0
    prev_year_col = 0

    for s in semanas:
        d = ws.cell(row=row, column=column+idx, value=semanas[idx])
        d.fill = bg
        d.font = font
        d.border = borderfull
        d.alignment = alignment3
        if prev_year != anhos[semanas[idx]]:
            prev_year = anhos[semanas[idx]]
            d = ws.cell(row=row-2, column=column + idx, value=prev_year)
            d.fill = bg
            d.font = font
            d.alignment = alignment4
            if prev_year_col == 0:
                prev_year_col = column + idx
            else:
                ws.merge_cells(start_row=row-2, start_column=prev_year_col, end_row=row - 2, end_column=column+ idx -1)
                prev_year_col = column + idx

        if prev_month != meses[semanas[idx]]:
            prev_month = meses[semanas[idx]]
            d = ws.cell(row=row-1, column=column + idx, value=prev_month)
            d.fill = bg
            d.font = font
            d.border = borderfull
            d.alignment = alignment4
            if prev_month_col == 0:
                prev_month_col = column + idx
            else:
                ws.merge_cells(start_row=row-1, start_column=prev_month_col, end_row=row - 1, end_column=column+ idx -1)
                prev_month_col = column + idx
        idx += 1
    ws.merge_cells(start_row=row - 2, start_column=prev_year_col, end_row=row - 2, end_column=column + idx - 1)
    ws.merge_cells(start_row=row - 1, start_column=prev_month_col, end_row=row - 1, end_column=column + idx - 1)

    for aux_col in range(4,column+idx):
        d = ws.cell(row=row - 2, column=aux_col)
        d.border = borderbottom

    estados = OrderedDict()
    estados['descarga'] = 'Descarga en Parque'
    estados['premontaje'] = 'Pre-montaje'
    estados['montaje'] = 'Montaje'
    estados['puestaenmarcha'] = 'Puesta en marcha'
    #estados = ['descarga','premontaje','montaje','puestaenmarcha']
    estados_db = {}
    estados_db['descarga'] = EstadoFU.objects.get(idx=1)
    estados_db['premontaje'] = EstadoFU.objects.get(idx=2)
    estados_db['montaje'] = EstadoFU.objects.get(idx=3)
    estados_db['puestaenmarcha'] = EstadoFU.objects.get(idx=4)

    column = 2
    row = 4

    planExist = False
    if Plan.objects.filter(parque=parque).count()>0:
        planExist = True
    if Contractual.objects.filter(parque=parque).count()>0:
        planExist = True

    for e, titulo in estados.iteritems():
        componentes = getComponentesbyState(parque_eolico,e)
        if len(componentes) > 0:
            d = ws.cell(row=row, column=1, value=titulo)
            d.alignment = alignment1
            d.fill = bg
            d.font = font
            d.border = borderfull
            first_row = row
            first = True
            for idx, nombre in componentes.iteritems():
                d=ws.cell(row=row, column=column , value=nombre)
                d.alignment = alignment2
                d.fill = bg
                d.font = font
                ws.merge_cells(start_row=row, start_column=column, end_row=row + 1, end_column=column)
                if first:
                    d.border = bordertop
                    d = ws.cell(row=row + 1, column=column)
                    d.border = borderside
                    first = False
                else:
                    d.border = borderside
                    d = ws.cell(row=row+1, column=column)
                    d.border = borderside

                d2 = ws.cell(row=row, column=column+1, value='Contractual')
                d2.border = bordertop
                d2.alignment = alignment3
                d2 = ws.cell(row=row+1, column=column + 1, value='Plan')
                d2.border = borderbottom
                d2.alignment = alignment3
                columna_aux = 4
                if planExist :
                    for semana in semanas:
                        d_str = anhos[semana] + "-W" + semana + "-0"
                        fecha_aux = datetime.strptime(d_str, "%Y-W%W-%w")
                        try:
                            plan = Plan.objects.get(parque=parque,
                                                    componente__nombre=nombre,
                                                    estado=estados_db[e],
                                                    fecha=fecha_aux)
                            ws.cell(row=row+1, column=columna_aux, value=plan.no_aerogeneradores)
                        except Plan.DoesNotExist:
                            pass
                        try:
                            contractual = Contractual.objects.get(parque=parque,
                                                                  componente__nombre=nombre,
                                                                  estado=estados_db[e],
                                                                  fecha=fecha_aux)
                            ws.cell(row=row , column=columna_aux, value=contractual.no_aerogeneradores)
                        except Contractual.DoesNotExist:
                            pass
                        columna_aux += 1

                row += 2
            d.border = borderbottom
            ws.merge_cells(start_row=first_row, start_column=1, end_row=row-1, end_column=1)


    target_stream = StringIO.StringIO()
    wb.save(target_stream)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PlantilaPlanificacion.xlsx"'
    target_stream.flush()
    ret_excel = target_stream.getvalue()
    target_stream.close()
    response.write(ret_excel)
    return response


# 2: El componente-estado ya fue ingresado
# 1: El componente-estado está bloqueado por prerequisitos
# 0: El componente-estado puede ser ingresado
def getComponenteStatus(registros, idx, componente, miembros):
    aux = registros.filter(componente=componente, estado__idx=idx)
    if aux.count() > 0:
        return 2

    if idx == 1: # Estado descarga
        return 0
    elif idx == 2: # Estado pre-montaje
        estado = EstadoFU.objects.get(idx=2)
        # Si el componentes tiene estados anteriores al pre-montaje
        if miembros.filter(componente=componente, estado__idx__lt=2).count() > 0:
            aux2 = registros.filter(componente=componente, estado__idx=1)
            #
            if aux2.count() > 0:
                return 0
            else:
                return 1
        else:
            return 0

    elif idx == 3: # Estado montaje
        estado = EstadoFU.objects.get(idx=3)
        # Si ya está ingreasado Izado Cable HV, entonces el resto se puede ingresar
        if registros.filter(estado__idx=3, componente__nombre='Izado Cable HV').count() > 0:
            return 0
        #if componente.estados.all().filter(idx__lt=3).count() > 0:
        # Si el componente tiene estados anteriores al montaje
        prerequisitos = miembros.filter(componente=componente, estado__idx__lt=3)
        if prerequisitos.count() > 0:
            test = True
            for prerequisito in prerequisitos:
                # Se pide omitir pre-montaje de los prerequisitos
                if prerequisito.estado.idx != 2:
                    test = test and registros.filter(componente=componente, estado__idx=prerequisito.estado.idx).exists()
            # Si ya se han registrado los prerequisitos del componente en descarga o pre-montaje
            if test:
                c_aux = miembros.get(componente=componente, estado=estado)
                c_ids = []
                for m in miembros.filter(estado = estado, orden__lt=c_aux.orden, orden__gt=0):
                    c_ids.append(m.componente.id)
                aux3 = registros.filter(componente_id__in=c_ids,estado__idx=3)
                if aux3.count() == len(c_ids):
                    return 0
                else:
                    return 1
            else:
                return 1
        # Si el componente no tiene estados anteriores al montaje
        else:
            c_aux = miembros.get(componente=componente, estado=estado)
            c_ids = []
            for m in miembros.filter(orden__lt=c_aux.orden, orden__gt=0, estado=estado):
                c_ids.append(m.componente.id)
            aux3 = registros.filter(componente_id__in=c_ids, estado__idx=3)
            if aux3.count() == len(c_ids):
                return 0
            else:
                return 1
    elif idx == 4: # Estado puesta en marcha
        estado_montaje = EstadoFU.objects.get(idx=3)
        c_ids = []
        # Si ya está ingresado el izado, entonces se puede cualquiera
        if registros.filter(estado__idx=3,componente__nombre='Izado Cable HV').count() > 0:
            return 0
        for m in miembros.filter(estado=estado_montaje, orden__gt=0):
            c_ids.append(m.componente.id)
        aux3 = registros.filter(componente_id__in=c_ids, estado__idx=3)
        if aux3.count() == len(c_ids):
            return 0
        else:
            return 1
    else:
        return 1

    return 1


@login_required(login_url='ingresar')
def ingreso(request,slug,slug_ag):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    aerogenerador = get_object_or_404(Aerogenerador, parque=parque, slug=slug_ag)
    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        configuracion = None

    try:
        parque_eolico = ParqueEolico.objects.get(parque=parque)
    except ParqueEolico.DoesNotExist:
        parque_eolico = ParqueEolico(parque=parque)
        parque_eolico.save()

    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo= u'Ingreso Registros Aerogenerador ' + aerogenerador.nombre
    contenido.subtitulo= u'Parque Eólico ' +parque.nombre
    contenido.menu = ['menu-fu', 'menu2-ingreso-'+str(aerogenerador.idx)]

    if request.method == 'POST':
        registro = None
        if 'formDescarga' in request.POST:
            if not request.user.has_perm('fu.add_registros'):
                raise PermissionDenied
            formDescarga= RegistroDescargaForm(request.POST)
            if formDescarga.is_valid():
                registro = formDescarga.save(commit=False)
                estado = EstadoFU.objects.get(idx=1)
            else:
                messages.add_message(request, messages.ERROR, 'Registro no pudo realizarse')
        elif 'formPremontaje' in request.POST:
            if not request.user.has_perm('fu.add_registros'):
                raise PermissionDenied
            form = RegistroForm(request.POST)
            if form.is_valid():
                registro = form.save(commit=False)
                estado = EstadoFU.objects.get(idx=2)
            else:
                messages.add_message(request, messages.ERROR, 'Registro no pudo realizarse')
        elif 'formMontaje' in request.POST:
            if not request.user.has_perm('fu.add_registros'):
                raise PermissionDenied
            form = RegistroForm(request.POST)
            if form.is_valid():
                registro = form.save(commit=False)
                estado = EstadoFU.objects.get(idx=3)
            else:
                messages.add_message(request, messages.ERROR, 'Registro no pudo realizarse')
        elif 'formPuestaenmarcha' in request.POST:
            if not request.user.has_perm('fu.add_registros'):
                raise PermissionDenied
            form = RegistroForm(request.POST)
            if form.is_valid():
                registro = form.save(commit=False)
                estado = EstadoFU.objects.get(idx=4)
            else:
                messages.add_message(request, messages.ERROR, 'Registro no pudo realizarse')
        elif 'delete' in request.POST:
            componente_id = int(request.POST['del_id'])
            estado_id = int(request.POST['del_estado_id'])
            c = Componente.objects.get(id=componente_id)
            reg = Registros.objects.get(parque=parque, aerogenerador=aerogenerador,componente=c, estado__idx=estado_id)
            if not (request.user.has_perm('fu.delete_registros') or reg.created_by == request.user):
                raise PermissionDenied
            log_msg = "Se elimina registro para parque " + parque.nombre + \
                      " - Aerogenerador " + reg.aerogenerador.nombre
            log = Log(texto=log_msg, tipo=3, user=request.user)
            log.save()
            reg.delete()
            messages.add_message(request, messages.SUCCESS, 'Registro eliminado con éxito!')
        if registro is not None:
            registro.parque = parque
            registro.aerogenerador = aerogenerador
            componente_id = int(request.POST['id'])
            componente = Componente.objects.get(id=componente_id)
            registro.componente = componente
            registro.estado = estado
            registro.created_by = request.user
            registro.save()
            log_msg = "Se agrega registro para parque " + parque.nombre + \
                      " - Aerogenerador " + registro.aerogenerador.nombre
            log = Log(texto=log_msg, tipo=1, user=request.user)
            log.save()
            messages.add_message(request, messages.SUCCESS, 'Registro realizado con éxito!')

    componentes = OrderedDict()

    componentes['Descarga en Parque'] = {}
    componentes['Pre-montaje'] = {}
    componentes['Montaje'] = {}
    componentes['Puesta en marcha'] = {}

    registros = Registros.objects.filter(parque=parque, aerogenerador=aerogenerador)
    miembros = Membership.objects.filter(parque_eolico = parque_eolico)

    componentes['Descarga en Parque']['objetos'] = []
    # filtro = 'relacionesfu__orden_descarga'
    # karws = {filtro + '__gt': 0}
    estado = EstadoFU.objects.get(nombre='Descarga')
    for m in miembros.filter(estado=estado).order_by('orden'):
        aux = getComponenteStatus(registros, 1, m.componente, miembros)
        objeto = {}
        color = 'bg-yellow-crusta'
        if aux == 2:
            color = 'bg-green-meadow'
            try:
                reg = registros.get(componente=m.componente, estado__idx=1)
            except Registros.MultipleObjectsReturned:
                regs = registros.filter(componente=m.componente, estado__idx=1)
                prev_reg = None
                for reg in regs:
                    if prev_reg is not None:
                        prev_reg.delete()
                    prev_reg = reg
                    logger.debug('ERROR: Doble registro!')
                    logger.debug(reg)
            objeto['tooltip'] = reg.fecha.strftime("%d/%m/%Y") + '<br>' + reg.no_serie
            objeto['created_by'] = reg.created_by.id
        objeto['componente'] = m.componente
        objeto['color'] = color
        objeto['status'] = aux
        objeto['estado'] = 1
        componentes['Descarga en Parque']['objetos'].append(objeto)

    componentes['Pre-montaje']['objetos'] = []
    # filtro = 'relacionesfu__orden_montaje'
    # karws = {filtro + '__gt': 0}
    estado = EstadoFU.objects.get(nombre='Pre-Montaje')
    for m in miembros.filter(estado=estado).order_by('orden'):
        aux = getComponenteStatus(registros, 2, m.componente, miembros)
        color = 'bg-grey-salt'
        objeto = {}
        if aux == 2:
            color = 'bg-green-meadow'
            reg = registros.get(componente=m.componente, estado__idx=2)
            objeto['tooltip'] = reg.fecha.strftime("%d/%m/%Y")
            objeto['created_by'] = reg.created_by.id
        elif aux == 0:
            color = 'bg-yellow-crusta'
        objeto['componente'] = m.componente
        objeto['color'] = color
        objeto['status'] = aux
        objeto['estado'] = 2
        componentes['Pre-montaje']['objetos'].append(objeto)

    componentes['Montaje']['objetos']  = []
    # filtro = 'relacionesfu__orden_montaje'
    # karws = {filtro + '__gt': 0}
    estado = EstadoFU.objects.get(nombre='Montaje')
    for m in miembros.filter(estado=estado).order_by('orden'):
        aux = getComponenteStatus(registros, 3, m.componente, miembros)
        color = 'bg-grey-salt'
        objeto = {}
        if aux == 2:
            color = 'bg-green-meadow'
            try:
                reg = registros.get(componente=m.componente, estado__idx=3)
            except MultipleObjectsReturned:
                regs = registros.filter(componente=m.componente, estado__idx=3)
                prev_reg = None
                for reg in regs:
                    if prev_reg is not None:
                        prev_reg.delete()
                    prev_reg = reg
                    logger.debug('ERROR: Doble registro!')
                    logger.debug(reg)

            objeto['tooltip'] = reg.fecha.strftime("%d/%m/%Y")
            objeto['created_by'] = reg.created_by.id
        elif aux == 0:
            color = 'bg-yellow-crusta'
        objeto['componente'] = m.componente
        objeto['color'] = color
        objeto['status'] = aux
        objeto['estado'] = 3
        componentes['Montaje']['objetos'].append(objeto)

    componentes['Puesta en marcha']['objetos']  = []
    # filtro = 'relacionesfu__orden_puestaenmarcha'
    # karws = {filtro + '__gt': 0}
    estado = EstadoFU.objects.get(nombre='Puesta en marcha')
    for m in miembros.filter(estado=estado).order_by('orden'):
        aux = getComponenteStatus(registros, 4, m.componente, miembros)
        color = 'bg-grey-salt'
        objeto = {}
        if aux == 2:
            color = 'bg-green-meadow'
            reg = registros.get(componente=m.componente, estado__idx=4)
            objeto['tooltip'] = reg.fecha.strftime("%d/%m/%Y")
            objeto['created_by'] = reg.created_by.id
        elif aux == 0:
            color = 'bg-yellow-crusta'
        objeto['componente'] = m.componente
        objeto['color'] = color
        objeto['status'] = aux
        objeto['estado'] = 4
        componentes['Puesta en marcha']['objetos'].append(objeto)


    componentes['Descarga en Parque']['id'] = 'descarga'
    componentes['Pre-montaje']['id'] = 'premontaje'
    componentes['Montaje']['id'] = 'montaje'
    componentes['Puesta en marcha']['id'] = 'puestaenmarcha'

    componentes['Descarga en Parque']['icon'] = 'fa-map-marker'
    componentes['Pre-montaje']['icon'] = 'fa-cog'
    componentes['Montaje']['icon'] = 'fa-cogs'
    componentes['Puesta en marcha']['icon'] = 'fa-thumbs-o-up'


    formDescarga = RegistroDescargaForm()
    form = RegistroForm()

    pos = {}
    estado = EstadoFU.objects.get(idx=3)
    registros_aux = Registros.objects.filter(parque=parque,
                                         aerogenerador=aerogenerador,
                                         estado=estado)
    if registros_aux.count() > 0:
        reg_ids = []
        for r in registros_aux:
            reg_ids.append(r.componente.id)
        # rel = RelacionesFU.objects.filter(componentes_parque=componentes_parque,
        #                                   componente__in=reg_ids).order_by('-orden_montaje')

        estado_montaje = EstadoFU.objects.get(nombre='Montaje')
        members = Membership.objects.filter(parque_eolico=parque_eolico,
                                            componente__in=reg_ids,
                                            estado = estado_montaje).order_by('-orden')
        for m in members:
            path = os.path.join(settings.BASE_DIR, 'static/common/images/ag')
            filename = path + '/' + m.componente.nombre + '.png'
            if os.path.isfile(unicode(filename).encode('utf-8')):
                [width, top, left] = get_image_data_ingreso(os.path.basename(filename))
                pos['width'] = width
                pos['top'] = top
                pos['left'] = left
                pos['img'] = 'common/images/ag/' + m.componente.nombre + '.png'
                break


    return TemplateResponse(request, 'fu/ingreso.html',
                  {'cont': contenido,
                   'parque': parque,
                   'aerogeneradores': aerogeneradores,
                   'form': form,
                   'componentes': componentes,
                   'aerogenerador': aerogenerador,
                   'formDescarga': formDescarga,
                   'registros':registros.order_by('fecha'),
                   'pos':pos,
                   })


def get_image_data_ingreso(filename):
    if filename == 'Pala 3.png':
        width = 75
        top = 10
        left = 4  # OK
    elif filename == 'T1.png':
        width = 7
        top = 136
        left = 43  # OK
    elif filename == 'T2.png':
        width = 10
        top = 110
        left = 42  # OK
    elif filename == 'T3.png':
        width = 11
        top = 64
        left = 41  # OK
    elif filename == 'Nacelle.png':
        width = 11
        top = 63
        left = 42  # OK
    elif filename == 'Buje.png':
        width = 14
        top = 60
        left = 38  # OK
    elif filename == 'Pala 1.png':
        width = 20
        top = 3
        left = 36  # OK
    elif filename == 'Pala 2.png':
        width = 73
        top = 55
        left = 2
    elif filename == 'Rotor.png':
        width = 75
        top = 10
        left = 4  # OK

    return [width,top,left]


@login_required(login_url='ingresar')
def paradas(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Listado de Paradas '
    contenido.subtitulo='Parque '+ parque.nombre
    contenido.menu = ['menu-fu', 'menu2-paradas']

    paradas = Paradas.objects.filter(parque=parque)

    if request.method == 'POST':
        if 'del_id' in request.POST:
            id = int(request.POST['del_id'])
            parada = Paradas.objects.get(id=id)
            if not (request.user.has_perm('fu.delete_paradas') or request.user == parada.created_by):
                raise PermissionDenied
            log_msg = "Se elimina parada para parque " + parque.nombre + \
                      " - Aerogenerador - " + parada.aerogenerador.nombre
            log = Log(texto=log_msg, tipo=3, user=request.user)
            log.save()
            parada.delete()
            messages.add_message(request, messages.SUCCESS, 'Registro eliminado con éxito!')
        else:
            messages.add_message(request, messages.ERROR, 'Error al eliminar registro')

    return TemplateResponse(request, 'fu/paradas.html',
        {'cont': contenido,
            'parque': parque,
            'aerogeneradores': aerogeneradores,
            'paradas': paradas,
        })


@login_required(login_url='ingresar')
@permission_required('fu.add_paradas', raise_exception=True)
def add_paradas(request,slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Listado de Paradas '
    contenido.subtitulo='Parque '+ parque.nombre
    contenido.menu = ['menu-fu', 'menu2-paradas']

    form = None

    if request.method == 'POST':
        form = ParadasForm(request.POST)
        if form.is_valid():
            if not request.user.has_perm('fu.add_paradas'):
                raise PermissionDenied
            parada = form.save(commit=False)
            parada.parque = parque
            parada.created_by = request.user
            parada.save()
            log_msg = "Se agrega parada para parque " + parque.nombre + \
                      " - Aerogenerador - " + parada.aerogenerador.nombre
            log = Log(texto=log_msg, tipo=1, user=request.user)
            log.save()
            #parada.delete()
            messages.add_message(request, messages.SUCCESS, 'Registro agregado con éxito!')
            return HttpResponseRedirect(reverse('fu:paradas', args=[parque.slug]))
        else:
            messages.add_message(request, messages.ERROR, 'Error al agregar el registro')


    if form is None:
        form = ParadasForm(initial={'parque':parque})
    back_url = reverse('fu:paradas', args=[parque.slug])
    return TemplateResponse(request, 'fu/agregarParada.html',
        {'cont': contenido,
            'parque': parque,
            'form': form,
            'back_url':back_url,
            'aerogeneradores': aerogeneradores,
        })


@login_required(login_url='ingresar')
def edit_paradas(request,slug,id):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    parada = get_object_or_404(Paradas, id=id)
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Listado de Paradas '
    contenido.subtitulo='Parque '+ parque.nombre
    contenido.menu = ['menu-fu', 'menu2-paradas']

    form = None

    if request.method == 'POST':
        form = ParadasForm(request.POST, instance=parada)
        if form.is_valid():
            if not (request.user.has_perm('fu.edit_paradas') or request.user == parada.created_by):
                raise PermissionDenied
            parada = form.save()
            log_msg = "Se edita parada para parque " + parque.nombre + \
                      " - Aerogenerador - " + parada.aerogenerador.nombre
            log = Log(texto=log_msg, tipo=2, user=request.user)
            log.save()
            messages.add_message(request, messages.SUCCESS, 'Registro editado con éxito!')
            return HttpResponseRedirect(reverse('fu:paradas', args=[parque.slug]))
        else:
            messages.add_message(request, messages.SUCCESS, 'Error al editar el registro')

    if form is None:
        form = ParadasForm(instance=parada)
    back_url = reverse('fu:paradas', args=[parque.slug])
    edit_parada = parada
    return TemplateResponse(request, 'fu/agregarParada.html',
        {'cont': contenido,
            'parque': parque,
            'form': form,
            'aerogeneradores': aerogeneradores,
            'back_url':back_url,
            'edit_parada': edit_parada,
        })


def nested_dict(n, type):
    if n == 1:
        return defaultdict(type)
    else:
        return defaultdict(lambda: nested_dict(n-1, type))


def dataPlanificacion(parque):
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    datos = nested_dict(2, int)
    filas = Node("Filas")
    columnas = Node("Columnas")
    estados = EstadoFU.objects.filter(idx__gt=0).order_by('id')
    # filtros = {}
    # filtros[2] = 'relacionesfu__orden_descarga'
    # filtros[4] = 'relacionesfu__orden_montaje'
    # filtros[5] = 'relacionesfu__orden_puestaenmarcha'

    datos2 = OrderedDict()
    configuracion = ConfiguracionFU.objects.get(parque=parque)
    final = configuracion.fecha_final
    last_anho = 0
    last_mes = 0
    r = Resolver('name')
    fila = 0

    for e in estados:
        miembros = Membership.objects.filter(parque_eolico=parque_eolico, estado = e)
        if miembros.count() > 0:
            n_fila = Node(e.nombre, filas)
            for m in miembros:
                n2_fila = Node(m.componente.nombre, n_fila)
                n_contractual = Node('Contractual', n2_fila)
                n_plan = Node('Plan', n2_fila)
                aux = configuracion.fecha_inicio
                columna = 0
                datos2[fila]=OrderedDict()
                datos2[fila +1 ] = OrderedDict()
                filtro_contractual = Contractual.objects.filter(parque=parque,
                                                              componente=m.componente,
                                                              estado=e)
                filtro_plan = Plan.objects.filter(parque=parque,
                                                componente=m.componente,
                                                estado=e)
                while aux < final:
                    if last_anho != aux.year:
                        try:
                            n_anho =r.get(columnas, str(aux.year))
                        except:
                            n_anho = Node(str(aux.year), columnas)
                        last_anho = aux.year
                    if last_mes != aux.month:
                        try:
                            n_mes = r.get(n_anho, meses_espanol[str(aux.month)])
                        except:
                            n_mes = Node(meses_espanol[str(aux.month)], n_anho)
                        last_mes = aux.month
                    try:
                        n_semana = r.get(n_mes,str(aux.isocalendar()[1]))
                    except:
                        n_semana = Node(str(aux.isocalendar()[1]), n_mes)
                    d_str = str(aux.isocalendar()[0]) + "-W" + str(aux.isocalendar()[1]) + "-0"
                    fecha_aux = datetime.strptime(d_str, "%Y-W%W-%w")
                    try:
                        plan = filtro_plan.get(fecha=fecha_aux)
                        #datos[n_plan][n_semana] = plan.no_aerogeneradores
                        datos2[fila+1][columna] = plan.no_aerogeneradores
                    except Plan.DoesNotExist:
                        pass
                    try:
                        contractual = filtro_contractual.get(fecha=fecha_aux)
                        #datos[n_contractual][n_semana] = contractual.no_aerogeneradores
                        datos2[fila][columna] = contractual.no_aerogeneradores
                    except Contractual.DoesNotExist:
                        pass

                    aux = aux + relativedelta.relativedelta(weeks=1)
                    columna += 1
                fila += 2
    return [filas,columnas,datos2]


def printFilasHeader(ws,node,row,column,deep,alignment1=None):
    deep +=1
    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    alignments = {}
    if alignment1 is None:
        alignments[1] = Alignment(horizontal="center", vertical="center", text_rotation=90)
    else:
        alignments[1] = alignment1
    alignments[2] = Alignment(vertical="center")
    alignments[3] = Alignment(horizontal="center")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)

    for level in node.children:
        d = ws.cell(row=row, column=column+deep-1, value=level.name)
        d.alignment = alignments[deep]
        d.fill = bg
        d.font = font

        first_row = row

        if len(level.children) > 0:
            row = printFilasHeader(ws,level,row,column,deep)
        else:
            #d.border = borderfull
            row += 1
        #d.border = borderbottom
        if (first_row != (row - 1)):
            ws.merge_cells(start_row=first_row, start_column=column+deep-1, end_row=row - 1, end_column=column+deep-1)
        else:
            d.border = borderfull

    return row


def printColumnasHeader(ws,node,row,column,deep):
    deep +=1
    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    alignments = {}
    alignments[1] = Alignment(horizontal="center", vertical="center", text_rotation=90)
    alignments[2] = Alignment(vertical="center")
    alignments[3] = Alignment(horizontal="center")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)

    for level in node.children:
        d = ws.cell(row=row+deep-1, column=column, value=level.name)
        d.fill = bg
        d.font = font
        d.border = borderfull
        d.alignment = alignments[3]
        start_column = column
        if len(level.children) > 0:
            column = printColumnasHeader(ws,level,row,column,deep)
        else:
            column += 1
        if (start_column != (column-1)):
            ws.merge_cells(start_row=row + deep -1, start_column=start_column, end_row=row+deep -1,
                       end_column=column  - 1)
    return column


def getHojas(node, hojas):
    if len(node.children) == 0:
        hojas.append(node)
    else:
        for level in node.children:
            getHojas(level,hojas)


def printDataExcel(ws,datos, initial_row,initial_col,alignment = None, column_width = None):
    thin = Side(border_style="thin", color="000000")
    borderbottom = Border(left=thin, right=thin, bottom=thin)

    max_row = 0
    max_col = 0
    for fila, aux in datos.iteritems():
        if fila > max_row:
            max_row = fila
        for columna, value in aux.iteritems():
            if columna > max_col:
                max_col = columna
            d = ws.cell(row=initial_row + fila, column=initial_col + columna, value=value)
            if alignment is not None:
                d.alignment = alignment

    rango = get_column_letter(initial_col) + str(initial_row) + ':' + \
            get_column_letter(initial_col + max_col) + str(initial_row + max_row)

    rows = ws[rango]
    for row in rows:
        for cell in row:
            cell.border = borderbottom

    if column_width is not None:
        for columna in range(initial_col,initial_col+max_col+1):
            ws.column_dimensions[get_column_letter(columna)].width = column_width


def planificacionExcel(parque, wb):
    [filas,columnas,datos] = dataPlanificacion(parque)
    if 'Planificacion'  in wb.sheetnames:
        ws = wb['Planificacion']
    else:
        ws = wb.create_sheet("Planificacion")

    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)
    borderbottom = Border(left=thin, right=thin, bottom=thin)
    borderside = Border(left=thin, right=thin)
    alignment3 = Alignment(horizontal="center")

    row = 3

    d = ws.cell(row=row, column=1, value='Actividad')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=2, value='Componente')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=3, value='Semana')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    ws.column_dimensions["A"].width = 11.0
    ws.column_dimensions["B"].width = 25.0
    ws.column_dimensions["C"].width = 11.0

    printColumnasHeader(ws,columnas,1,4,0)
    printFilasHeader(ws, filas, 4, 1, 0)
    for cell_range in ws.merged_cell_ranges:
        rows = ws[cell_range]
        for cell in rows[0]:
            cell.border = borderside
            cell.border = borderbottom
        for cell in rows[-1]:
            cell.border = borderside
            cell.border = borderbottom

    printDataExcel(ws,datos,4,4)

    return wb


def addUniqueNodo(nombre, padre, html=None):
    r = Resolver('name')
    try:
        nodo = r.get(padre, nombre)
    except:
        nodo = Node(nombre, padre)
    if html is not None:
        if nombre not in html:
            html.append(nombre)

    return nodo


def dataSeriesfechas(parque, html=None):
    datos = OrderedDict()
    filas = Node("Filas")
    columnas = Node("Columnas")
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    estados = EstadoFU.objects.filter(idx__gt=0).order_by('id')

    configuracion = ConfiguracionFU.objects.get(parque=parque)
    r = Resolver('name')
    ags = Aerogenerador.objects.filter(parque=parque,nombre__istartswith='WTG').order_by('idx')
    fila = 0
    filas_html = []
    columnas_html = []
    total_filas = Membership.objects.filter(parque_eolico=parque_eolico, estado__id=2).count() * 2
    total_filas += Membership.objects.filter(parque_eolico=parque_eolico).exclude(estado__id=2).count()
    suma_filas = [0] * (total_filas + 1)

    for e in estados:
        miembros = Membership.objects.filter(parque_eolico=parque_eolico, estado = e)
        if miembros.count() > 0:
            logger.debug('Estado: ' + e.nombre)
            n_fila = Node(e.nombre, filas)
            for m in miembros:
                logger.debug(' - Componente: ' + m.componente.nombre)
                filtro_registros = Registros.objects.filter(parque=parque,componente=m.componente,estado=e)
                n2_fila = Node(m.componente.nombre, n_fila)
                datos[fila] = OrderedDict()
                n_fecha = Node('Fecha', n2_fila)
                filas_html.append(
                    {'Estado': e.nombre, 'Componente': m.componente.nombre, 'Item': 'Fecha', 'Fila': fila})
                if e.id == 2:
                    n_serie = Node('Serie', n2_fila)
                    filas_html.append(
                        {'Estado': e.nombre, 'Componente': m.componente.nombre, 'Item': 'Serie', 'Fila': fila+1})
                    datos[fila+1] = OrderedDict()
                columna = 0
                for ag in ags:
                    addUniqueNodo(ag.nombre,columnas,columnas_html)
                    try:
                        r = filtro_registros.get(aerogenerador = ag)
                        datos[fila][columna] = r.fecha.strftime('%d-%m-%Y')
                        try:
                            suma_filas[fila] += 1
                        except:
                            suma_filas[fila] = 1
                        if e.id == 2:
                            datos[fila+1][columna] = r.no_serie
                            try:
                                suma_filas[fila+1] += 1
                            except:
                                suma_filas[fila+1] = 1
                    except Registros.DoesNotExist:
                        pass
                    columna += 1
                addUniqueNodo('Total', columnas, columnas_html)
                #datos[fila][columna] = suma_filas[fila]
                try:
                    datos[fila][columna] = suma_filas[fila]
                except IndexError:
                    print('Error')
                if e.id == 2:
                    datos[fila+1][columna] = suma_filas[fila+1]
                    fila += 2
                else:
                    fila += 1

    if html is None:
        return [filas,columnas,datos]
    else:
        return [filas_html,columnas_html,datos]


def seriesfechasExcel(parque,wb):
    [filas, columnas, datos] = dataSeriesfechas(parque)
    sheet_name = 'Número de series y fecha'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)
    borderbottom = Border(left=thin, right=thin, bottom=thin)
    borderside = Border(left=thin, right=thin)
    alignment3 = Alignment(horizontal="center")

    row = 2

    d = ws.cell(row=row, column=1, value='Actividad')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=2, value='Componente')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=3, value='Elemento')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    ws.column_dimensions["A"].width = 11.0
    ws.column_dimensions["B"].width = 25.0
    ws.column_dimensions["C"].width = 11.0

    printColumnasHeader(ws, columnas, 2, 4, 0)
    printFilasHeader(ws, filas, 3, 1, 0)
    for cell_range in ws.merged_cell_ranges:
        rows = ws[cell_range]
        if len(rows) > 1:
            for row in rows:
                row[0].border = borderside
                row[0].border = borderbottom
        else:
            for cell in rows[0]:
                cell.border = borderside
                cell.border = borderbottom

    printDataExcel(ws, datos, 3, 4,alignment3,14)

    if not parque.logo_excel:
        parque.create_excel_logo()
        parque.save()

    img1 = Image(parque.logo_excel)

    img2 = Image(os.path.join(settings.BASE_DIR, 'static/common/images/saroenlogo-excel.png'))
    ws.add_image(img1, 'A1')
    ws.add_image(img2, 'K1')

    return wb


def dataSeguimiento(parque,t, html = None):
    datos = OrderedDict()
    filas = Node("Filas")
    columnas = Node("Columnas")
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    estados = EstadoFU.objects.filter(idx__gt=0).order_by('id')
    # componentes_parque = ComponentesParque.objects.get(parque=parque).componentes.all()
    # filtros = {}
    # filtros[2] = 'relacionesfu__orden_descarga'
    # filtros[4] = 'relacionesfu__orden_montaje'
    # filtros[5] = 'relacionesfu__orden_puestaenmarcha'
    configuracion = ConfiguracionFU.objects.get(parque=parque)
    final = configuracion.fecha_final

    ags = Aerogenerador.objects.filter(parque=parque, nombre__istartswith='WTG').order_by('idx')
    fila = 0

    anho = t.year
    semana = t.isocalendar()[1]
    d = str(t.isocalendar()[0]) + '-W' + str(semana)
    fecha_calculo = datetime.strptime(d + '-0', "%Y-W%W-%w")

    filas_html = []
    columnas_html = []

    for e in estados:
        miembros = Membership.objects.filter(parque_eolico=parque_eolico, estado=e).order_by('orden')
        if miembros.count() > 0:
            n_fila = Node(e.nombre, filas)
            for m in miembros:
                n2_fila = Node(m.componente.nombre, n_fila)
                filas_html.append(
                    {'Estado': e.nombre, 'Componente': m.componente.nombre, 'Fila': fila})
                filtro_registros = Registros.objects.filter(parque=parque, componente=m.componente, estado=e)
                columna = 0
                datos[fila] = OrderedDict()
                # Estadisticas primero
                addUniqueNodo('Total', columnas, columnas_html)
                datos[fila][columna] = parque.no_aerogeneradores
                columna += 1
                addUniqueNodo('Contractual', columnas, columnas_html)
                datos[fila][columna] = getContractual(parque, m.componente, e, fecha_calculo)
                columna += 1
                addUniqueNodo('Plan', columnas, columnas_html)
                datos[fila][columna] = getPlan(parque, m.componente, e, fecha_calculo)
                columna += 1
                addUniqueNodo('Real', columnas, columnas_html)
                datos[fila][columna] = getReal(parque, m.componente, e, t)
                columna += 1

                for ag in ags:
                    addUniqueNodo(ag.nombre,columnas, columnas_html)
                    try:
                        r = filtro_registros.get(aerogenerador = ag)
                        datos[fila][columna] = 'X'
                    except Registros.DoesNotExist:
                        pass
                    columna += 1
                fila += 1

    if html is None:
        return [filas,columnas,datos]
    else:
        return [filas_html,columnas_html,datos]


def seguimientoExcel(parque,wb,t):
    [filas, columnas, datos] = dataSeguimiento(parque,t)
    sheet_name = 'Seguimiento'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)
    borderbottom = Border(left=thin, right=thin, bottom=thin)
    borderside = Border(left=thin, right=thin)
    alignment3 = Alignment(horizontal="center")

    row = 2

    d = ws.cell(row=row, column=1, value='Actividad')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3
    d = ws.cell(row=row, column=2, value='Componente')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    ws.column_dimensions["A"].width = 11.0
    ws.column_dimensions["B"].width = 25.0

    printColumnasHeader(ws, columnas, 2, 3, 0)
    printFilasHeader(ws, filas, 3, 1, 0)
    for cell_range in ws.merged_cell_ranges:
        rows = ws[cell_range]
        if len(rows) > 1:
            for row in rows:
                row[0].border = borderside
                row[0].border = borderbottom
        else:
            for cell in rows[0]:
                cell.border = borderside
                cell.border = borderbottom

    printDataExcel(ws, datos, 3, 3, alignment3, 6)
    ws.column_dimensions[get_column_letter(4)].width = 9


    img1 = Image(parque.logo_excel)
    img2 = Image(os.path.join(settings.BASE_DIR,'static/common/images/saroenlogo-excel.png'))
    ws.add_image(img1, 'A1')
    ws.add_image(img2, 'K1')

    return wb


def dataTasaMontaje(parque,t, html = None):
    datos = OrderedDict()
    suma_columna = OrderedDict()
    suma_fila = OrderedDict()
    filas = Node("Filas")
    columnas = Node("Columnas")
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    # componentes_parque = ComponentesParque.objects.get(parque=parque).componentes.all()
    # filtros = {}
    # filtros[2] = 'relacionesfu__orden_descarga'
    # filtros[4] = 'relacionesfu__orden_montaje'
    # filtros[5] = 'relacionesfu__orden_puestaenmarcha'

    configuracion = ConfiguracionFU.objects.get(parque=parque)

    d_str = str(t.isocalendar()[0]) + "-W" + str(t.isocalendar()[1]) + "-0"
    fecha_informe = datetime.strptime(d_str, "%Y-W%W-%w").date()

    final = configuracion.fecha_final

    if fecha_informe > final:
        final = fecha_informe

    last_anho = 0
    last_mes = 0
    fila = 0

    aux = configuracion.fecha_inicio

    e = EstadoFU.objects.get(nombre='Montaje')
    e_descarga = EstadoFU.objects.get(nombre='Descarga')
    #n_componentes = componentes_parque.filter(relacionesfu__orden_montaje__gt=0, relacionesfu__orden_descarga__gt=0).count()
    n_componentes = aerogeneradoresAMontar(parque_eolico)

    [proyeccion,ultima_fecha] = calcularProyeccion(parque_eolico, t.isocalendar()[0], t.isocalendar()[1])

    filas_html = []
    columnas_html = []

    while aux < final:
        if last_anho != aux.year:
            n_anho = addUniqueNodo(str(aux.year),filas)
            last_anho = aux.year
        if last_mes != aux.month:
            n_mes = addUniqueNodo(meses_espanol[str(aux.month)], n_anho)
            last_mes = aux.month
        n_semana = addUniqueNodo(str(aux.isocalendar()[1]), n_mes)
        datos[fila] = OrderedDict()
        suma_fila[fila] = float(0)
        d_str = str(aux.isocalendar()[0]) + "-W" + str(aux.isocalendar()[1]) + "-0"
        fecha_aux = datetime.strptime(d_str, "%Y-W%W-%w")

        #karws = {filtros[e.id] + '__gt': 0, filtros[e.id] + '__lte': n_componentes}
        #componentes = componentes_parque.filter(**karws).order_by(filtros[e.id])
        miembros = Membership.objects.filter(parque_eolico=parque_eolico,
                                             estado=e,
                                             orden__range=(0, n_componentes))
        columna = 0
        if miembros.count() > 0:
            for m in miembros:
                if aux <= fecha_informe:
                    n_comp = addUniqueNodo(m.componente.nombre, columnas,columnas_html)
                    value = getReal(parque, m.componente, e, fecha_aux)

                    if fila > 0 :
                        datos[fila][columna] = value - suma_columna[columna]
                        suma_columna[columna] = value
                    else:
                        suma_columna[columna] = 0
                        datos[fila][columna] = value

                    suma_fila[fila] += datos[fila][columna]

                columna += 1

        n_comp = addUniqueNodo('Total Semana',columnas, columnas_html)
        if aux <= fecha_informe:
            datos[fila][columna] = round(float(suma_fila[fila])/ float(miembros.count()),2)
        columna += 1

        n_comp = addUniqueNodo('Real', columnas, columnas_html)
        if aux <= fecha_informe:
            if fila > 0:
                datos[fila][columna] = datos[fila][columna-1] + datos[fila-1][columna]
            else:
                datos[fila][columna] = datos[fila][columna-1]
        columna += 1

        n_comp = addUniqueNodo('Contractual', columnas,columnas_html)
        datos[fila][columna] = getContractual(parque, m.componente, e, fecha_aux)
        columna += 1

        n_comp = addUniqueNodo('Plan', columnas, columnas_html)
        datos[fila][columna] = getPlan(parque, m.componente, e, fecha_aux)
        columna += 1

        n_comp = addUniqueNodo('Proyección', columnas, columnas_html)

        try:
            datos[fila][columna] = round((proyeccion[fila]*parque.no_aerogeneradores)/100, 2)
        except:
            if ( len(proyeccion) > 0 ):
                datos[fila][columna] = round((proyeccion[-1]*parque.no_aerogeneradores)/100, 2)
            else:
                datos[fila][columna] = 0
        columna += 1

        filas_html.append(
            {'Anho': str(aux.isocalendar()[0]), 'Mes': meses_espanol[str(aux.month)], 'Sem': str(aux.isocalendar()[1]),
             'Fila': fila})

        aux = aux + relativedelta.relativedelta(weeks=1)
        fila += 1

    if html is None:
        return [filas,columnas,datos]
    else:
        return [filas_html,columnas_html,datos]


def tasaMontajeExcel(parque,wb,t):
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    [filas, columnas, datos] = dataTasaMontaje(parque, t)
    #componentes_parque = ComponentesParque.objects.get(parque=parque).componentes.all()
    miembros = Membership.objects.filter(parque_eolico=parque_eolico)
    n_componentes = aerogeneradoresAMontar(parque_eolico)

    sheet_name = 'Tasa de montaje'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)
    borderbottom = Border(left=thin, right=thin, bottom=thin)
    borderside = Border(left=thin, right=thin)
    alignment3 = Alignment(horizontal="center")

    row = 2

    d = ws.cell(row=row, column=1, value='Año')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    d = ws.cell(row=row, column=2, value='Mes')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    d = ws.cell(row=row, column=3, value='Sem')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8

    printColumnasHeader(ws, columnas, 2, 4, 0)
    printFilasHeader(ws, filas, 3, 1, 0)
    for cell_range in ws.merged_cell_ranges:
        rows = ws[cell_range]
        if len(rows) > 1:
            for row in rows:
                row[0].border = borderside
                row[0].border = borderbottom
        else:
            for cell in rows[0]:
                cell.border = borderside
                cell.border = borderbottom

    printDataExcel(ws, datos, 3, 4, alignment3, 10)

    for col in range(12,12+6):
        ws.column_dimensions[get_column_letter(col)].width = 12

    last_row = datos.keys()[-1] + 3 + 1
    last_col = datos[0].keys()[-1] +4 -5 +1
    d = ws.cell(row=last_row, column=3, value='Total')
    d.fill = bg
    d.font = font
    d.alignment = alignment3
    for col in range(4,last_col+1):
        suma = '=SUM(' + get_column_letter(col) + str(5) +':'+ get_column_letter(col) + str(last_row-1) +')'
        d = ws.cell(row=last_row, column=col, value=suma)
        d.alignment = alignment3
        d.border = borderbottom

    img1 = Image(parque.logo_excel)
    img2 = Image(os.path.join(settings.BASE_DIR, 'static/common/images/saroenlogo-excel.png'))
    ws.add_image(img1, 'A1')
    ws.add_image(img2, 'M1')

    return wb


def dataListadoParadas(parque, t, html=None):
    datos = OrderedDict()
    filas = Node("Filas")
    columnas = Node("Columnas")

    configuracion = ConfiguracionFU.objects.get(parque=parque)

    fila = 0

    paradas = Paradas.objects.filter(fecha_inicio__lte=t, parque=parque).order_by('fecha_inicio')

    filas_html = []
    columnas_html = []

    for parada in paradas:
        datos[fila] = OrderedDict()
        n_row = addUniqueNodo(str(parada.id), filas)
        n_col = addUniqueNodo('WTG', columnas, columnas_html)
        datos[fila][0] = parada.aerogenerador.nombre
        n_col = addUniqueNodo('Componente', columnas, columnas_html)
        if parada.componente is not None:
            datos[fila][1] = parada.componente.nombre
        else:
            datos[fila][1] = ''
        n_col = addUniqueNodo('Trabajo', columnas, columnas_html)
        datos[fila][2] = parada.trabajo.nombre
        n_col = addUniqueNodo('Hora Paralización', columnas, columnas_html)
        datos[fila][3] = localtime(parada.fecha_inicio).strftime('%d-%m-%Y %H:%M')
        n_col = addUniqueNodo('Hora fin Paralización', columnas, columnas_html)
        datos[fila][4] = localtime(parada.fecha_final).strftime('%d-%m-%Y %H:%M')
        n_col = addUniqueNodo('Duración Paralización (h)', columnas, columnas_html)
        datos[fila][5] = parada.duracion
        n_col = addUniqueNodo('Motivo', columnas, columnas_html)
        datos[fila][6] = parada.motivo
        n_col = addUniqueNodo('Grúa', columnas, columnas_html)
        if parada.grua is not None:
            datos[fila][7] = parada.grua.nombre
        else:
            datos[fila][7] = ''
        n_col = addUniqueNodo('Observaciones', columnas, columnas_html)
        datos[fila][8] = parada.observaciones
        filas_html.append({'Item': str(parada.id), 'Fila': fila})
        fila += 1

    if html is None:
        return [filas,columnas,datos]
    else:
        return [filas_html,columnas_html,datos]


def listadoParadasExcel(parque, wb, t):
    [filas, columnas, datos] = dataListadoParadas(parque, t)
    sheet_name = 'Listado de paradas'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    bgColor = Color(rgb="283861")
    bg = PatternFill(patternType='solid', fgColor=bgColor)
    font = Font(color="FFFFFF")
    thin = Side(border_style="thin", color="000000")
    borderfull = Border(top=thin, left=thin, right=thin, bottom=thin)
    borderbottom = Border(left=thin, right=thin, bottom=thin)
    borderside = Border(left=thin, right=thin)
    alignment3 = Alignment(horizontal="center")

    row = 2

    d = ws.cell(row=row, column=1, value='Item')
    d.fill = bg
    d.font = font
    d.border = borderfull
    d.alignment = alignment3

    ws.column_dimensions["A"].width = 6

    printColumnasHeader(ws, columnas, 2, 2, 0)
    printFilasHeader(ws, filas, 3, 1, 0, Alignment(horizontal="center"))
    for cell_range in ws.merged_cell_ranges:
        rows = ws[cell_range]
        if len(rows) > 1:
            for row in rows:
                row[0].border = borderside
                row[0].border = borderbottom
        else:
            for cell in rows[0]:
                cell.border = borderside
                cell.border = borderbottom

    printDataExcel(ws, datos, 3, 2, alignment3, 6)

    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 12.2
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 17.6
    ws.column_dimensions["F"].width = 17.6
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 25
    ws.column_dimensions["I"].width = 10.6
    ws.column_dimensions["J"].width = 50

    img1 = Image(parque.logo_excel)
    img2 = Image(os.path.join(settings.BASE_DIR, 'static/common/images/saroenlogo-excel.png'))
    ws.add_image(img1, 'A1')
    ws.add_image(img2, 'H1')

    return wb


def graficosFUExcel(parque,wb):
    # filtros = {}
    # filtros[1] = 'relacionesfu__orden_descarga'
    # filtros[3] = 'relacionesfu__orden_montaje'
    # filtros[4] = 'relacionesfu__orden_puestaenmarcha'

    sheet_name = 'Graficas'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    ws2 = wb['Seguimiento']
    ws.sheet_view.zoomScale = 80
    font_test = TextFont(typeface='Calibri')

    #componentes_parque = ComponentesParque.objects.get(parque=parque).componentes.all()
    parque_eolico = ParqueEolico.objects.get(parque=parque)
    miembros = Membership.objects.filter(parque_eolico=parque_eolico)
    estados = EstadoFU.objects.filter(idx__gt=0).order_by('id')
    n_graficos = 0
    cuenta_anterior = 0
    for e in estados:
        # if e.idx in filtros:
        #     karws = {filtros[e.idx] + '__gt': 0}
        #     componentes = componentes_parque.filter(**karws).order_by(filtros[e.idx])
        miembros = Membership.objects.filter(parque_eolico=parque_eolico, estado=e).order_by('orden')
        if miembros.count() > 0:
            cuenta = miembros.distinct().count()
            c = BarChart()
            c.style =  2
            data1 =  Reference(ws2, min_col=3, max_col=6 ,min_row=3+cuenta_anterior, max_row=3 + cuenta_anterior + cuenta-1)
            categs = Reference(ws2, min_col=2, max_col=2, min_row=3+cuenta_anterior, max_row=3 + cuenta_anterior + cuenta-1)
            cuenta_anterior += cuenta
            # 2018-07-03 Se solicita que no se muestre este gráfico
            if e.nombre != 'Pre-montaje':
                c.add_data(data1,titles_from_data=False)

                bgColor = []
                bgColor.append(ColorChoice(srgbClr="a5a5a5"))
                bgColor.append(ColorChoice(srgbClr="2E75B6"))
                bgColor.append(ColorChoice(srgbClr="ED7D31"))
                bgColor.append(ColorChoice(srgbClr="70AD47"))
                legend_names = []
                legend_names.append('C2')
                legend_names.append('D2')
                legend_names.append('E2')
                legend_names.append('F2')

                for i in range(0,4):
                    c.series[i].graphicalProperties.solidFill = bgColor[i]
                    c.series[i].dLbls = DataLabelList()
                    c.series[i].dLbls.showVal = True
                    cp = CharacterProperties(latin=font_test, sz=800)
                    c.series[i].dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                    title = u"{0}!{1}".format('Seguimiento', legend_names[i])
                    title = SeriesLabel(strRef=StrRef(title))
                    c.series[i].tx = title

                c.overlap = -20
                c.set_categories(categs)

                c.y_axis.scaling.min = 0
                c.height = 10.22
                c.width = 25.7
                if e.nombre == 'Descarga':
                    c.title = 'Descarga en Parque'
                else:
                    c.title = e.nombre
                c.y_axis.title = 'Nº de Componentes'
                c.legend.position = 'b'
                lnAxis = LineProperties(noFill=True,w =0)
                c.graphical_properties = GraphicalProperties(ln=lnAxis)
                c.y_axis.spPr = GraphicalProperties(ln=lnAxis)
                c.plot_area.graphicalProperties = GraphicalProperties(ln=lnAxis)

                cp = CharacterProperties(latin=font_test, sz=900)
                c.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                c.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                c.y_axis.title.tx.rich.p[0].r.rPr = cp

                lnFont = LineProperties(prstDash = "solid", solidFill="244185")
                cp2 =  CharacterProperties(latin=font_test, sz=1400, solidFill = "24C885")

                c.title.tx.rich.p[0].r.rPr = cp2
                c.title.tx.rich.p[0].endParaRpr = cp2
                c.title.tx.rich.p[0].r.rPr.solidFill.RGB="24C885"
                #c.title.txPr = GraphicalProperties(solidFill="244185")
                fila_grafico = 3 + 20 * (n_graficos)
                pos = "B" + str(fila_grafico)
                ws.add_chart(c, pos)
                #ws.cell(row=fila_grafico,column=1,value=c.style)
                n_graficos += 1

    # Gráfica de tasa de montaje
    ws3 = wb['Tasa de montaje']
    fila = 3
    valor = ws3.cell(row=fila,column = 3).value
    while valor is not None:
        fila += 1
        valor = ws3.cell(row=fila, column=3).value

    e = EstadoFU.objects.get(idx=3)
    n_componentes = aerogeneradoresAMontar(parque_eolico)

    cols = n_componentes
    data_col = 3 + cols + 2
    c = LineChart()
    data2 = Reference(ws3, min_col=data_col, max_col=data_col + 4 -1, min_row=2, max_row=fila - 2)
    categs = Reference(ws3, min_col=3, min_row=3, max_row=fila -2)
    c.add_data(data2, titles_from_data=True)
    c.set_categories(categs)
    c.height = 10.22
    c.width = 25.7

    c.title = 'Avance de Izado'

    c.y_axis.title = 'Nº de Aerogeneradores montados'
    c.legend.position = 'b'
    lnAxis = LineProperties(noFill=True, w=0)
    c.graphical_properties = GraphicalProperties(ln=lnAxis)
    c.y_axis.spPr = GraphicalProperties(ln=lnAxis)
    c.plot_area.graphicalProperties = GraphicalProperties(ln=lnAxis)

    bgColor = []
    bgColor.append(ColorChoice(srgbClr="70AD47"))
    bgColor.append(ColorChoice(srgbClr="5B9BD5"))
    bgColor.append(ColorChoice(srgbClr="ED7D31"))
    bgColor.append(ColorChoice(srgbClr="70A977"))
    for i in range(0, 4):
        c.series[i].graphicalProperties.solidFill = bgColor[i]
        c.series[i].graphicalProperties.line.solidFill = bgColor[i]
        c.series[i].graphicalProperties.line.width = 20050
        cp = CharacterProperties(latin=font_test, sz=800)

    c.series[3].graphicalProperties.line.dashStyle = "dot"

    cp = CharacterProperties(latin=font_test, sz=900)
    c.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    c.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    c.y_axis.title.tx.rich.p[0].r.rPr = cp

    cp2 = CharacterProperties(latin=font_test, sz=1400, solidFill="24C885")
    c.title.tx.rich.p[0].r.rPr = cp2
    #c.x_axis.crossAx = 1
    pos = "R3"
    ws.add_chart(c, pos)

    img1 = Image(parque.logo_excel)
    img2 = Image(os.path.join(settings.BASE_DIR, 'static/common/images/saroenlogo-excel.png'))
    ws.add_image(img1, 'A1')
    ws.add_image(img2, 'N1')

    return wb


def reporteExcelFU(parque,t,nombre):
    if parque.excel_fu:
        wb = load_workbook(parque.excel_fu.file.name)
    else:
        wb = Workbook()
    #wb = planificacionExcel(parque, wb)
    wb = seriesfechasExcel(parque, wb)
    wb = seguimientoExcel(parque,wb,t)
    wb = tasaMontajeExcel(parque,wb,t)
    wb = listadoParadasExcel(parque,wb,t)
    wb = graficosFUExcel(parque,wb)

    target_stream = StringIO.StringIO()
    wb.save(target_stream)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="' + nombre +'.xlsx"'
    target_stream.flush()
    ret_excel = target_stream.getvalue()
    target_stream.close()
    response.write(ret_excel)

    return response


def reportePdfFU(parque,t):
    with open(os.path.join(settings.BASE_DIR, 'static/common/images/saroenlogo.png'), "rb") as image_file:
        logo_saroen = base64.b64encode(image_file.read())

    # data Series y Fechas
    resultados = []
    columnas=[]
    datos=[]
    [resultados,columnas,datos] = dataSeriesfechas(parque,True)
    titulo = 'Número de serie y fechas'
    nombre = ''
    paginas = {}
    ag_per_page = 8
    page = 0

    for ag in range(0,parque.no_aerogeneradores+1):
        if (ag) % ag_per_page == 0:
            page +=1
            paginas[page]=[]
        paginas[page].append(ag)

    # data Seguimiento
    filas_seguimiento = []
    col_seguimiento=[]
    seguimiento=[]
    [filas_seguimiento, col_seguimiento, seguimiento] = dataSeguimiento(parque, t, True)
    paginas_seguimiento = {}
    ag_per_page = 14
    page = 0

    for ag in range(0, len(col_seguimiento)):
        if (ag) % ag_per_page == 0:
            page += 1
            paginas_seguimiento[page] = []
        paginas_seguimiento[page].append(ag)

    # data Tasa de montaje
    filas_tasamontaje = []
    col_tasamontaje = []
    tasamontaje = []
    [filas_tasamontaje, col_tasamontaje, tasamontaje] = dataTasaMontaje(parque, t, True)
    paginas_tasamontaje = {}
    col_per_page = 14
    page = 0

    for col in range(0, len(col_tasamontaje)):
        if (col) % col_per_page == 0:
            page += 1
            paginas_tasamontaje[page] = []
        paginas_tasamontaje[page].append(col)

    # data Listado de paradas
    filas_paradas = []
    col_paradas = []
    paradas = []
    [filas_paradas, col_paradas, paradas] = dataListadoParadas(parque, t, True)


    pdf = render_to_pdf('fu/reporteFU.html',
                        {'pagesize': 'LETTER landscape',
                         'title': 'Reporte Punchlist',
                         'resultados': resultados,
                         'datos': datos,
                         'parque': parque,
                         'titulo': titulo,
                         'logo_saroen': logo_saroen,
                         'fecha': t,
                         'nombre': nombre,
                         'paginas': paginas,
                         'columnas': columnas,
                         'filas_seguimiento': filas_seguimiento,
                         'col_seguimiento': col_seguimiento,
                         'seguimiento': seguimiento,
                         'paginas_seguimiento': paginas_seguimiento,
                         'filas_tasamontaje': filas_tasamontaje,
                         'col_tasamontaje': col_tasamontaje,
                         'tasamontaje': tasamontaje,
                         'paginas_tasamontaje': paginas_tasamontaje,
                         'filas_paradas': filas_paradas,
                         'col_paradas': col_paradas,
                         'paradas': paradas
                         })
    return pdf


@login_required(login_url='ingresar')
def reportes(request, slug):
    parque = get_object_or_404(ParqueSolar, slug=slug)
    aerogeneradores = Aerogenerador.objects.filter(parque=parque).order_by('idx')
    contenido=ContenidoContainer()
    contenido.user=request.user
    contenido.titulo=u'Reportes Follow Up '
    contenido.subtitulo='Parque '+ parque.nombre
    contenido.menu = ['menu-fu', 'menu2-reportes']

    try:
        configuracion = ConfiguracionFU.objects.get(parque=parque)
    except ConfiguracionFU.DoesNotExist:
        configuracion = None


    form = None

    if request.method == 'POST':
        form = ReporteForm(request.POST)
        if form.is_valid():
            t = form.cleaned_data['fecha']
            if form.cleaned_data['nombre_archivo'] != '':
                nombre = form.cleaned_data['nombre_archivo']
            else:
                nombre = 'ReporteFU'
            if 'pdf' in request.POST:
                pdf =  reportePdfFU(parque,t)
                respuesta = HttpResponse(pdf,content_type='application/pdf')
                respuesta['Content-Disposition'] = 'attachment; filename="' + nombre +'.pdf"'
                return respuesta
            if 'excel' in request.POST:
                return reporteExcelFU(parque,t,nombre)
    if form is None:
        form = ReporteForm()

    return TemplateResponse(request, 'fu/reportes.html',
        {'cont': contenido,
            'parque': parque,
            'form': form,
            'aerogeneradores': aerogeneradores,
            'configuracion': configuracion
        })