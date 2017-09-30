# -*- coding: utf-8 -*-
import os, django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ControlObras.settings")
django.setup()

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl import load_workbook
from datetime import datetime

from vista.models import ParqueSolar, Aerogenerador
from fu.models import ComponentesParque, Componente, EstadoFU, Registros


if __name__ == '__main__':

    parques = ParqueSolar.objects.all()
    for parque in parques:
        parque.prev_no_aerogeneradores = parque.no_aerogeneradores
        Aerogenerador.objects.filter(parque=parque, idx__gt=parque.no_aerogeneradores).delete()

